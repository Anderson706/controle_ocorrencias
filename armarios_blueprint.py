# armarios_blueprint.py — Gestão de Armários integrado ao CCTV Control Panel
from flask import (
    Blueprint, render_template, request, redirect,
    url_for, flash, session, jsonify, send_file
)
from functools import wraps
from datetime import datetime
from io import BytesIO
import base64

armarios_bp = Blueprint('armarios', __name__, template_folder='templates')

_db    = None
Armario             = None
ArmarioChaveReserva = None
ArmarioHistorico    = None


def _tempo_decorrido(dt):
    delta     = datetime.now() - dt
    total_seg = int(delta.total_seconds())
    if total_seg < 0:
        return "agora"
    dias  = total_seg // 86400
    horas = (total_seg % 86400) // 3600
    mins  = (total_seg % 3600) // 60
    partes = []
    if dias:  partes.append(f"{dias}d")
    if horas: partes.append(f"{horas}h")
    if mins or not partes: partes.append(f"{mins}min")
    return " ".join(partes)


# ── Inicialização ─────────────────────────────────────────────────────────────
def setup_armarios(db):
    global _db, Armario, ArmarioChaveReserva, ArmarioHistorico
    _db = db

    class _Armario(db.Model):
        """Armário físico do site. Atribuição fixa ao colaborador."""
        __tablename__ = "armario"
        id                   = db.Column(db.Integer, db.Sequence('armario_id_seq', start=1), primary_key=True)
        numero               = db.Column(db.String(50),  nullable=False)
        bloco                = db.Column(db.String(100),  nullable=True)
        site                 = db.Column(db.String(100),  nullable=False)
        status               = db.Column(db.String(20),   nullable=False, default='LIVRE')  # LIVRE | OCUPADO
        colaborador_nome     = db.Column(db.String(150),  nullable=True)
        colaborador_cpf      = db.Column(db.String(50),   nullable=True)
        atribuido_em         = db.Column(db.DateTime,     nullable=True)
        atribuido_por        = db.Column(db.String(150),  nullable=True)   # operador que fez a atribuição
        assinatura_atribuicao = db.Column(db.Text,        nullable=True)   # base64 PNG do canvas
        criado_em            = db.Column(db.DateTime,     default=datetime.now)
        ativo                = db.Column(db.Boolean,      default=True)

    class _ArmarioChaveReserva(db.Model):
        """Registro de retirada/devolução da chave reserva de um armário."""
        __tablename__ = "armario_chave_reserva"
        id                  = db.Column(db.Integer, db.Sequence('arm_chave_id_seq', start=1), primary_key=True)
        armario_id          = db.Column(db.Integer,     nullable=False)
        retirado_por_nome   = db.Column(db.String(150), nullable=False)
        retirado_por_cpf    = db.Column(db.String(50),  nullable=False)
        responsavel_entrega = db.Column(db.String(150), nullable=False)
        usuario_id          = db.Column(db.Integer,     nullable=False)
        site                = db.Column(db.String(100), nullable=False)
        data_retirada       = db.Column(db.DateTime,    nullable=False, default=datetime.now)
        data_devolucao      = db.Column(db.DateTime,    nullable=True)
        status              = db.Column(db.String(20),  nullable=False, default='RETIRADA')  # RETIRADA | DEVOLVIDA
        assinatura          = db.Column(db.Text,        nullable=True)  # base64 PNG do canvas

    class _ArmarioHistorico(db.Model):
        """Auditoria de todas as operações nos armários."""
        __tablename__ = "armario_historico"
        id               = db.Column(db.Integer, db.Sequence('arm_hist_id_seq', start=1), primary_key=True)
        armario_id       = db.Column(db.Integer,     nullable=False)
        armario_numero   = db.Column(db.String(50),  nullable=True)
        bloco            = db.Column(db.String(100), nullable=True)
        site             = db.Column(db.String(100), nullable=False)
        evento           = db.Column(db.String(50),  nullable=False)  # ATRIBUIÇÃO | LIBERAÇÃO | CHAVE RETIRADA | CHAVE DEVOLVIDA
        colaborador_nome = db.Column(db.String(150), nullable=True)
        colaborador_cpf  = db.Column(db.String(50),  nullable=True)
        operador         = db.Column(db.String(150), nullable=True)
        data_evento      = db.Column(db.DateTime,    nullable=False, default=datetime.now)
        observacao       = db.Column(db.String(300), nullable=True)

    Armario             = _Armario
    ArmarioChaveReserva = _ArmarioChaveReserva
    ArmarioHistorico    = _ArmarioHistorico

    import sys
    mod = sys.modules[__name__]
    mod.Armario             = _Armario
    mod.ArmarioChaveReserva = _ArmarioChaveReserva
    mod.ArmarioHistorico    = _ArmarioHistorico
    mod._db                 = db


# ── Decorators ────────────────────────────────────────────────────────────────
def _login_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return f(*a, **kw)
    return dec

def _is_privileged():
    """Retorna True se o usuário for ADMIN (acesso cross-site total no módulo de armários).
    GESTOR usa site próprio; KEYUSER usa site próprio."""
    return (session.get("user_perfil") or "").upper() in ("ADMIN",)

def _is_can_manage():
    """Retorna True para perfis que podem gerenciar armários (inclui KEYUSER)."""
    return (session.get("user_perfil") or "").upper() in ("ADMIN", "GESTOR", "KEYUSER")

def _admin_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if not _is_can_manage():
            flash("Acesso restrito a administradores, gestores e key users.", "danger")
            return redirect(url_for("armarios.painel"))
        return f(*a, **kw)
    return dec


# ── Helpers ───────────────────────────────────────────────────────────────────
def _chave_ativa(armario_id):
    """Retorna o registro de chave reserva ativo para o armário, ou None."""
    return ArmarioChaveReserva.query.filter_by(
        armario_id=armario_id, status='RETIRADA'
    ).first()


def _reg_historico(armario_id, armario_numero, bloco, site, evento,
                   colaborador_nome=None, colaborador_cpf=None,
                   operador=None, observacao=None):
    """Registra um evento no histórico de armários. Silencioso em caso de erro."""
    try:
        h = ArmarioHistorico(
            armario_id       = armario_id,
            armario_numero   = armario_numero,
            bloco            = bloco,
            site             = site,
            evento           = evento,
            colaborador_nome = colaborador_nome,
            colaborador_cpf  = colaborador_cpf,
            operador         = operador,
            data_evento      = datetime.now(),
            observacao       = observacao,
        )
        _db.session.add(h)
        _db.session.commit()
    except Exception:
        _db.session.rollback()


# ═══════════════════════════════════════════════════════════════════════════════
# PAINEL PRINCIPAL
# ═══════════════════════════════════════════════════════════════════════════════

@armarios_bp.route("/")
@_login_required
def painel():
    site     = session.get("user_site", "")
    is_admin = _is_privileged()

    bloco_filtro  = request.args.get("bloco",  "")
    status_filtro = request.args.get("status", "")
    site_filtro   = request.args.get("site",   "") if is_admin else ""

    # ── Uma query para TODOS os armários (KPIs + blocos + lista) ────
    if is_admin:
        todos = (Armario.query
                 .filter_by(ativo=True)
                 .order_by(Armario.site, Armario.bloco, Armario.numero)
                 .all())
    else:
        todos = (Armario.query
                 .filter_by(site=site, ativo=True)
                 .order_by(Armario.bloco, Armario.numero)
                 .all())

    total    = len(todos)
    ocupados = sum(1 for a in todos if a.status == 'OCUPADO')
    livres   = sum(1 for a in todos if a.status == 'LIVRE')
    blocos   = sorted({a.bloco for a in todos if a.bloco})
    sites_lista = sorted({a.site for a in todos if a.site}) if is_admin else []

    # Filtragem em Python (evita segunda query ao banco)
    armarios = [a for a in todos
                if (not bloco_filtro  or a.bloco  == bloco_filtro)
                and (not status_filtro or a.status == status_filtro)
                and (not site_filtro  or a.site   == site_filtro)]

    # ── UMA query para todas as chaves reserva ativas (elimina N+1) ──────────
    arm_ids      = [a.id for a in todos]
    chaves_ativas_map = {}
    if arm_ids:
        rows = (ArmarioChaveReserva.query
                .filter(ArmarioChaveReserva.armario_id.in_(arm_ids))
                .filter_by(status='RETIRADA')
                .all())
        chaves_ativas_map = {r.armario_id: r for r in rows}

    agora      = datetime.now()
    chave_fora = 0
    for arm in todos:
        ch = chaves_ativas_map.get(arm.id)
        if ch:
            chave_fora += 1
    for arm in armarios:
        ch = chaves_ativas_map.get(arm.id)
        arm._chave_ativa  = ch
        if ch:
            arm._chave_tempo  = _tempo_decorrido(ch.data_retirada)
            arm._chave_alerta = (agora - ch.data_retirada).days >= 1
        else:
            arm._chave_tempo  = None
            arm._chave_alerta = False

    return render_template(
        "armarios/painel.html",
        armarios=armarios,
        blocos=blocos,
        bloco_filtro=bloco_filtro,
        status_filtro=status_filtro,
        site_filtro=site_filtro,
        sites_lista=sites_lista,
        site=site,
        is_admin=is_admin,
        can_manage=_is_can_manage(),
        total=total, ocupados=ocupados, livres=livres, chave_fora=chave_fora,
    )


# ── Detalhe JSON (modal) ──────────────────────────────────────────────────────
@armarios_bp.route("/detalhe/<int:arm_id>")
@_login_required
def detalhe(arm_id):
    arm  = Armario.query.get_or_404(arm_id)
    site = session.get("user_site", "")
    if arm.site != site and not _is_privileged():
        return jsonify({"erro": "Acesso negado"}), 403
    ch = _chave_ativa(arm.id)
    return jsonify({
        "id":               arm.id,
        "numero":           arm.numero,
        "bloco":            arm.bloco or "—",
        "status":           arm.status,
        "colaborador_nome": arm.colaborador_nome or "—",
        "colaborador_cpf":  arm.colaborador_cpf  or "—",
        "atribuido_em":     arm.atribuido_em.strftime("%d/%m/%Y") if arm.atribuido_em else "—",
        "chave_retirada":   bool(ch),
        "chave_retirado_por":   ch.retirado_por_nome if ch else None,
        "chave_cpf":            ch.retirado_por_cpf  if ch else None,
        "chave_entregue_por":   ch.responsavel_entrega if ch else None,
        "chave_data":           ch.data_retirada.strftime("%d/%m/%Y %H:%M") if ch else None,
        "chave_tempo":          _tempo_decorrido(ch.data_retirada) if ch else None,
        "chave_id":             ch.id if ch else None,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# CRIAR EM LOTE (ADMIN)
# ═══════════════════════════════════════════════════════════════════════════════

@armarios_bp.route("/criar-lote", methods=["POST"])
@_login_required
@_admin_required
def criar_lote():
    site       = session.get("user_site", "")
    prefixo    = request.form.get("prefixo", "ARM").strip().upper()
    bloco      = request.form.get("bloco", "").strip()
    num_inicio = request.form.get("num_inicio", type=int, default=1)
    num_fim    = request.form.get("num_fim",    type=int, default=10)
    digitos    = request.form.get("digitos",    type=int, default=3)

    if num_inicio > num_fim:
        flash("Número inicial deve ser menor que o final.", "warning")
        return redirect(url_for("armarios.painel"))
    if num_fim - num_inicio > 199:
        flash("Limite de 200 armários por lote.", "warning")
        return redirect(url_for("armarios.painel"))

    criados = 0
    for n in range(num_inicio, num_fim + 1):
        numero = f"{prefixo}-{str(n).zfill(digitos)}"
        # Pula se já existe esse número no site
        existe = Armario.query.filter_by(site=site, numero=numero, ativo=True).first()
        if existe:
            continue
        arm = Armario(numero=numero, bloco=bloco or None, site=site, status='LIVRE')
        _db.session.add(arm)
        criados += 1

    try:
        _db.session.commit()
        flash(f"{criados} armário(s) criado(s) com sucesso!", "success")
    except Exception as e:
        _db.session.rollback()
        flash(f"Erro ao criar armários: {e}", "danger")

    return redirect(url_for("armarios.painel"))


# ═══════════════════════════════════════════════════════════════════════════════
# ATRIBUIR / LIBERAR ARMÁRIO
# ═══════════════════════════════════════════════════════════════════════════════

@armarios_bp.route("/atribuir/<int:arm_id>", methods=["POST"])
@_login_required
def atribuir(arm_id):
    arm  = Armario.query.get_or_404(arm_id)
    nome = request.form.get("colaborador_nome", "").strip()
    cpf  = request.form.get("colaborador_cpf",  "").strip()

    if not nome or not cpf:
        flash("Informe o nome e CPF/matrícula do colaborador.", "warning")
        return redirect(url_for("armarios.painel"))

    # Cada CPF só pode ter um armário por site
    existente = Armario.query.filter_by(
        colaborador_cpf=cpf, site=arm.site, status='OCUPADO', ativo=True
    ).first()
    if existente and existente.id != arm_id:
        flash(f"Este CPF/matrícula já está atribuído ao armário {existente.numero}. "
              f"Libere-o antes de atribuir outro.", "danger")
        return redirect(url_for("armarios.painel"))

    assinatura   = request.form.get("assinatura") or None
    atribuido_por = session.get("user_nome", "")
    agora         = datetime.now()

    # Tentativa 1: salva todos os campos (inclusive novos)
    try:
        arm.colaborador_nome      = nome
        arm.colaborador_cpf       = cpf
        arm.status                = 'OCUPADO'
        arm.atribuido_em          = agora
        arm.atribuido_por         = atribuido_por
        arm.assinatura_atribuicao = assinatura
        _db.session.commit()
        _reg_historico(arm.id, arm.numero, arm.bloco, arm.site, 'ATRIBUIÇÃO',
                       colaborador_nome=nome, colaborador_cpf=cpf, operador=atribuido_por)
        flash(f"Armário {arm.numero} atribuído a {nome} com sucesso!", "success")
        return redirect(url_for("armarios.painel"))
    except Exception:
        _db.session.rollback()

    # Tentativa 2: fallback sem colunas que podem não existir ainda no Oracle
    _atrib_ok = False
    try:
        _db.session.execute(
            _db.text(
                "UPDATE ARMARIO SET COLABORADOR_NOME=:nome, COLABORADOR_CPF=:cpf, "
                "STATUS='OCUPADO', ATRIBUIDO_EM=:agora WHERE ID=:id"
            ),
            {"nome": nome, "cpf": cpf, "agora": agora, "id": arm_id},
        )
        _db.session.commit()
        flash(f"Armário {arm.numero} atribuído a {nome} com sucesso! "
              f"(Assinatura será disponível após reiniciar o sistema.)", "success")
        _atrib_ok = True
    except Exception as e2:
        _db.session.rollback()
        flash(f"Erro ao atribuir armário: {e2}", "danger")

    if _atrib_ok:
        _reg_historico(arm.id, arm.numero, arm.bloco, arm.site, 'ATRIBUIÇÃO',
                       colaborador_nome=nome, colaborador_cpf=cpf, operador=atribuido_por)

    return redirect(url_for("armarios.painel"))


@armarios_bp.route("/liberar/<int:arm_id>", methods=["POST"])
@_login_required
def liberar(arm_id):
    arm = Armario.query.get_or_404(arm_id)

    # Verifica se a chave reserva está fora
    if _chave_ativa(arm_id):
        flash("A chave reserva está retirada. Devolva-a antes de liberar o armário.", "danger")
        return redirect(url_for("armarios.painel"))

    try:
        colab_nome = arm.colaborador_nome
        colab_cpf  = arm.colaborador_cpf
        arm.colaborador_nome = None
        arm.colaborador_cpf  = None
        arm.status           = 'LIVRE'
        arm.atribuido_em     = None
        _db.session.commit()
        _reg_historico(arm.id, arm.numero, arm.bloco, arm.site, 'LIBERAÇÃO',
                       colaborador_nome=colab_nome, colaborador_cpf=colab_cpf,
                       operador=session.get("user_nome", ""))
        flash(f"Armário {arm.numero} liberado com sucesso!", "success")
    except Exception as e:
        _db.session.rollback()
        flash(f"Erro ao liberar armário: {e}", "danger")

    return redirect(url_for("armarios.painel"))


@armarios_bp.route("/excluir/<int:arm_id>", methods=["POST"])
@_login_required
@_admin_required
def excluir(arm_id):
    arm = Armario.query.get_or_404(arm_id)
    if arm.status == 'OCUPADO':
        flash("Libere o armário antes de excluí-lo.", "danger")
        return redirect(url_for("armarios.painel"))
    if _chave_ativa(arm_id):
        flash("Devolva a chave reserva antes de excluir.", "danger")
        return redirect(url_for("armarios.painel"))
    try:
        arm.ativo = False
        _db.session.commit()
        flash(f"Armário {arm.numero} removido.", "success")
    except Exception as e:
        _db.session.rollback()
        flash(f"Erro ao excluir: {e}", "danger")
    return redirect(url_for("armarios.painel"))


# ═══════════════════════════════════════════════════════════════════════════════
# EXCLUIR EM LOTE
# ═══════════════════════════════════════════════════════════════════════════════

@armarios_bp.route("/excluir-lote", methods=["POST"])
@_login_required
@_admin_required
def excluir_lote():
    ids      = request.form.getlist("ids")
    site     = session.get("user_site", "")
    is_admin = _is_privileged()
    if not ids:
        flash("Nenhum armário selecionado.", "warning")
        return redirect(url_for("armarios.painel"))
    removidos = ignorados = 0
    for arm_id in ids:
        q = Armario.query.filter_by(id=arm_id, ativo=True)
        if not is_admin:
            q = q.filter_by(site=site)
        arm = q.first()
        if not arm:
            continue
        if arm.status == 'OCUPADO':
            ignorados += 1
            continue
        if _chave_ativa(arm.id):
            ignorados += 1
            continue
        arm.ativo = False
        removidos += 1
    try:
        _db.session.commit()
        msg = f"{removidos} armário(s) removido(s)."
        if ignorados:
            msg += f" {ignorados} ignorado(s) (ocupados ou com chave retirada)."
        flash(msg, "success" if removidos else "warning")
    except Exception as e:
        _db.session.rollback()
        flash(f"Erro ao remover armários: {e}", "danger")
    return redirect(url_for("armarios.painel"))


# ═══════════════════════════════════════════════════════════════════════════════
# EDITAR COLABORADOR DE UM ARMÁRIO OCUPADO
# ═══════════════════════════════════════════════════════════════════════════════

@armarios_bp.route("/editar-colaborador/<int:arm_id>", methods=["POST"])
@_login_required
def editar_colaborador(arm_id):
    arm  = Armario.query.get_or_404(arm_id)
    site = session.get("user_site", "")

    if arm.site != site and not _is_privileged():
        flash("Acesso negado.", "danger")
        return redirect(url_for("armarios.painel"))

    nome = request.form.get("colaborador_nome", "").strip()
    cpf  = request.form.get("colaborador_cpf",  "").strip()

    if not nome or not cpf:
        flash("Informe o nome e CPF/matrícula do colaborador.", "warning")
        return redirect(url_for("armarios.painel"))

    # Verifica duplicidade de CPF em outro armário do mesmo site
    existente = Armario.query.filter_by(
        colaborador_cpf=cpf, site=arm.site, status='OCUPADO', ativo=True
    ).first()
    if existente and existente.id != arm_id:
        flash(f"Este CPF/matrícula já está atribuído ao armário {existente.numero}.", "danger")
        return redirect(url_for("armarios.painel"))

    try:
        arm.colaborador_nome = nome
        arm.colaborador_cpf  = cpf
        _db.session.commit()
        flash(f"Dados do colaborador do armário {arm.numero} atualizados!", "success")
    except Exception as e:
        _db.session.rollback()
        flash(f"Erro ao editar: {e}", "danger")

    return redirect(url_for("armarios.painel"))


# ═══════════════════════════════════════════════════════════════════════════════
# CHAVE RESERVA
# ═══════════════════════════════════════════════════════════════════════════════

@armarios_bp.route("/chave-reserva", methods=["GET", "POST"])
@_login_required
def chave_reserva():
    site       = session.get("user_site", "")
    usuario_id = session.get("user_id")

    if request.method == "POST":
        arm_id = request.form.get("armario_id", type=int)
        arm    = Armario.query.get(arm_id)

        if not arm or arm.site != site or not arm.ativo:
            flash("Armário inválido ou não pertence ao seu site.", "danger")
            return redirect(url_for("armarios.chave_reserva"))

        if _chave_ativa(arm_id):
            flash(f"A chave reserva do armário {arm.numero} já está retirada.", "danger")
            return redirect(url_for("armarios.chave_reserva"))

        cpf  = request.form.get("retirado_por_cpf", "").strip()
        nome = request.form.get("retirado_por_nome", "").strip()
        if not cpf or not nome:
            flash("Informe o nome e CPF/matrícula de quem está retirando a chave.", "warning")
            return redirect(url_for("armarios.chave_reserva"))

        try:
            reg = ArmarioChaveReserva(
                armario_id          = arm_id,
                retirado_por_nome   = nome,
                retirado_por_cpf    = cpf,
                responsavel_entrega = session.get("user_nome", ""),
                usuario_id          = usuario_id,
                site                = site,
                data_retirada       = datetime.now(),
                status              = 'RETIRADA',
                assinatura          = request.form.get("assinatura") or None,
            )
            _db.session.add(reg)
            _db.session.commit()
            _reg_historico(arm.id, arm.numero, arm.bloco, arm.site, 'CHAVE RETIRADA',
                           colaborador_nome=nome, colaborador_cpf=cpf,
                           operador=session.get("user_nome", ""))
            flash(f"Chave reserva do armário {arm.numero} registrada com sucesso!", "success")
        except Exception as e:
            _db.session.rollback()
            flash(f"Erro ao registrar retirada: {e}", "danger")

        return redirect(url_for("armarios.chave_reserva"))

    # Retiradas ativas
    ativas = (ArmarioChaveReserva.query
              .filter_by(site=site, status='RETIRADA')
              .order_by(ArmarioChaveReserva.data_retirada.asc())
              .all())
    for r in ativas:
        arm_r        = Armario.query.get(r.armario_id)
        r._numero    = arm_r.numero if arm_r else "?"
        r._bloco     = arm_r.bloco  if arm_r else "—"
        r._tempo     = _tempo_decorrido(r.data_retirada)
        r._alerta    = (datetime.now() - r.data_retirada).days >= 1

    # Armários do site para o select do formulário
    armarios_site = (Armario.query
                     .filter_by(site=site, ativo=True)
                     .order_by(Armario.bloco, Armario.numero)
                     .all())
    # Marca quais já têm chave fora
    for a in armarios_site:
        a._chave_fora = bool(_chave_ativa(a.id))

    return render_template(
        "armarios/chave_reserva.html",
        ativas=ativas,
        armarios_site=armarios_site,
        site=site,
        responsavel_nome=session.get("user_nome", ""),
    )


@armarios_bp.route("/chave-reserva/devolver/<int:reg_id>", methods=["POST"])
@_login_required
def devolver_chave_reserva(reg_id):
    reg = ArmarioChaveReserva.query.get_or_404(reg_id)
    if reg.status == 'DEVOLVIDA':
        flash("Chave já devolvida.", "warning")
        return redirect(url_for("armarios.chave_reserva"))
    try:
        arm_r              = Armario.query.get(reg.armario_id)
        reg.status         = 'DEVOLVIDA'
        reg.data_devolucao = datetime.now()
        _db.session.commit()
        _reg_historico(
            reg.armario_id,
            arm_r.numero if arm_r else None,
            arm_r.bloco  if arm_r else None,
            reg.site, 'CHAVE DEVOLVIDA',
            colaborador_nome=reg.retirado_por_nome,
            colaborador_cpf=reg.retirado_por_cpf,
            operador=session.get("user_nome", ""),
        )
        flash("Chave reserva devolvida com sucesso!", "success")
    except Exception as e:
        _db.session.rollback()
        flash(f"Erro ao registrar devolução: {e}", "danger")
    return redirect(url_for("armarios.chave_reserva"))


_TERMO_TEXTO = (
    "Recebi um armário pelo qual me responsabilizo, inclusive por sua boa conservação. "
    "Estou ciente de que devo avisar a Segurança Patrimonial quando não mais for utilizá-lo.\n\n"
    "Declaro estar ciente que devo utilizar esse armário apenas para a guarda de objetos pessoais "
    "de meu uso diário, sendo recomendada a não estocagem de valores, alimentos ou bebidas. "
    "Tenho conhecimento também que não poderei trocá-lo por outro sem autorização da área de "
    "Segurança Patrimonial. Concordo que o armário seja auditado periodicamente, na minha presença, "
    "pela Segurança Patrimonial/RH.\n\n"
    "Comprometo-me em trazer um cadeado de boa qualidade e uma cópia da chave de meu armário "
    "ficará sob a tutela da Segurança Patrimonial.\n\n"
    "Declaro ter ciência que em caso de esquecimento de minha chave deverei acionar um integrante "
    "da Segurança Patrimonial para devidas providências."
)


# ═══════════════════════════════════════════════════════════════════════════════
# TERMO DE RESPONSABILIDADE DE ARMÁRIO — PDF
# ═══════════════════════════════════════════════════════════════════════════════

@armarios_bp.route("/termo/<int:arm_id>")
@_login_required
def termo_atribuicao(arm_id):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, Image as RLImage,
    )

    arm = Armario.query.get_or_404(arm_id)
    site = session.get("user_site", "")
    if arm.site != site and not _is_privileged():
        flash("Acesso negado.", "danger")
        return redirect(url_for("armarios.painel"))

    output = BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm,
        topMargin=16*mm, bottomMargin=18*mm,
    )

    W    = A4[0] - 36*mm
    RED  = colors.HexColor("#d40511")
    DARK = colors.HexColor("#111827")
    GRAY = colors.HexColor("#6b7280")
    LGRY = colors.HexColor("#f3f4f6")

    h1   = ParagraphStyle("h1",  fontName="Helvetica-Bold", fontSize=18,
                           textColor=colors.white, alignment=TA_CENTER, leading=24)
    sub  = ParagraphStyle("sub", fontName="Helvetica",      fontSize=10,
                           textColor=colors.white, alignment=TA_CENTER, leading=14)
    lbl  = ParagraphStyle("lbl", fontName="Helvetica-Bold", fontSize=8,
                           textColor=GRAY, spaceAfter=2, leading=10)
    val  = ParagraphStyle("val", fontName="Helvetica-Bold", fontSize=13,
                           textColor=DARK, leading=16)
    sec  = ParagraphStyle("sec", fontName="Helvetica-Bold", fontSize=11,
                           textColor=RED, leading=14)
    body = ParagraphStyle("body", fontName="Helvetica", fontSize=10,
                           textColor=DARK, leading=15, alignment=TA_JUSTIFY, spaceAfter=8)
    rod  = ParagraphStyle("rod",  fontName="Helvetica", fontSize=8,
                           textColor=GRAY, alignment=TA_CENTER, leading=10)

    def campo(rotulo, texto):
        return [Paragraph(rotulo, lbl), Paragraph(str(texto) if texto else "—", val)]

    def linha_campos(pares, larguras):
        row0 = [campo(r, v)[0] for r, v in pares]
        row1 = [campo(r, v)[1] for r, v in pares]
        t = Table([row0, row1], colWidths=larguras)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), LGRY),
            ("TOPPADDING",    (0,0), (-1,-1), 8),
            ("BOTTOMPADDING", (0,0), (-1,-1), 8),
            ("LEFTPADDING",   (0,0), (-1,-1), 10),
            ("RIGHTPADDING",  (0,0), (-1,-1), 10),
        ]))
        return t

    def secao(titulo):
        t = Table([[Paragraph(titulo, sec)]], colWidths=[W])
        t.setStyle(TableStyle([
            ("LINEBELOW",     (0,0), (-1,-1), 1.5, RED),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ]))
        return t

    story = []

    # Cabeçalho
    cab = Table([[Paragraph("TERMO DE RESPONSABILIDADE DE ARMÁRIO", h1)]], colWidths=[W])
    cab.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), RED),
        ("TOPPADDING",    (0,0), (-1,-1), 14),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("LEFTPADDING",   (0,0), (-1,-1), 12),
        ("RIGHTPADDING",  (0,0), (-1,-1), 12),
        ("ROUNDEDCORNERS", [8]),
    ]))
    story.append(cab)

    inf = Table([[Paragraph(
        f"DHL SECURITY &nbsp;|&nbsp; {arm.site} &nbsp;|&nbsp; "
        f"Emitido em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", sub,
    )]], colWidths=[W])
    inf.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), DARK),
        ("TOPPADDING",    (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING",   (0,0), (-1,-1), 12),
        ("RIGHTPADDING",  (0,0), (-1,-1), 12),
    ]))
    story.append(inf)
    story.append(Spacer(1, 10*mm))

    # Dados do armário e colaborador
    story.append(secao("IDENTIFICAÇÃO DO ARMÁRIO E COLABORADOR"))
    story.append(Spacer(1, 4*mm))
    story.append(linha_campos(
        [("Nº DO ARMÁRIO", arm.numero),
         ("BLOCO", arm.bloco or "—"),
         ("SITE", arm.site)],
        [W*0.25, W*0.25, W*0.50],
    ))
    story.append(Spacer(1, 2))
    story.append(linha_campos(
        [("COLABORADOR RESPONSÁVEL", arm.colaborador_nome or "—"),
         ("CPF / MATRÍCULA", arm.colaborador_cpf or "—")],
        [W*0.55, W*0.45],
    ))
    story.append(Spacer(1, 2))
    story.append(linha_campos(
        [("DATA DE ATRIBUIÇÃO", arm.atribuido_em.strftime("%d/%m/%Y  %H:%M") if arm.atribuido_em else "—"),
         ("OPERADOR DE SEGURANÇA", arm.atribuido_por or "—")],
        [W*0.55, W*0.45],
    ))
    story.append(Spacer(1, 8*mm))

    # Texto do termo
    story.append(secao("DECLARAÇÃO DE RESPONSABILIDADE"))
    story.append(Spacer(1, 5*mm))

    # Caixa destacada com o texto
    termo_box = Table(
        [[Paragraph(_TERMO_TEXTO.replace("\n\n", "<br/><br/>"), body)]],
        colWidths=[W],
    )
    termo_box.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), colors.HexColor("#fffbf0")),
        ("BOX",           (0,0), (-1,-1), 0.8, colors.HexColor("#f5d66f")),
        ("TOPPADDING",    (0,0), (-1,-1), 14),
        ("BOTTOMPADDING", (0,0), (-1,-1), 14),
        ("LEFTPADDING",   (0,0), (-1,-1), 14),
        ("RIGHTPADDING",  (0,0), (-1,-1), 14),
    ]))
    story.append(termo_box)
    story.append(Spacer(1, 8*mm))

    # Assinatura
    story.append(HRFlowable(width=W, thickness=1, color=colors.HexColor("#e5e7eb")))
    story.append(Spacer(1, 6*mm))
    story.append(secao("ASSINATURA DO COLABORADOR"))
    story.append(Spacer(1, 5*mm))

    if arm.assinatura_atribuicao and arm.assinatura_atribuicao.startswith("data:image"):
        try:
            _, b64data = arm.assinatura_atribuicao.split(",", 1)
            img_buf = BytesIO(base64.b64decode(b64data))
            sig_img = RLImage(img_buf, width=120*mm, height=38*mm)
            sig_img.hAlign = "LEFT"
            sig_box = Table([[sig_img]], colWidths=[W])
            sig_box.setStyle(TableStyle([
                ("BACKGROUND",    (0,0), (-1,-1), colors.white),
                ("BOX",           (0,0), (-1,-1), 0.5, colors.HexColor("#d1d5db")),
                ("TOPPADDING",    (0,0), (-1,-1), 6),
                ("BOTTOMPADDING", (0,0), (-1,-1), 6),
                ("ALIGN",         (0,0), (-1,-1), "CENTER"),
            ]))
            story.append(sig_box)
        except Exception:
            story.append(_caixa_sig_vazia(W))
    else:
        story.append(_caixa_sig_vazia(W))

    story.append(Spacer(1, 10*mm))

    # Rodapé
    story.append(HRFlowable(width=W, thickness=0.5, color=colors.HexColor("#e5e7eb")))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        f"Armário {arm.numero} &nbsp;|&nbsp; DHL Security — {arm.site} &nbsp;|&nbsp; "
        f"Documento gerado automaticamente pelo sistema CCTV Control Panel",
        rod,
    ))

    doc.build(story)
    output.seek(0)
    fname = f"termo_armario_{arm.numero.replace('-','_')}_{arm.colaborador_cpf or 'sem_cpf'}.pdf"
    return send_file(output, as_attachment=True, download_name=fname,
                     mimetype="application/pdf")


# ═══════════════════════════════════════════════════════════════════════════════
# COMPROVANTE DE RETIRADA DE CHAVE RESERVA — PDF
# ═══════════════════════════════════════════════════════════════════════════════

@armarios_bp.route("/chave-reserva/comprovante/<int:reg_id>")
@_login_required
def comprovante_chave_reserva(reg_id):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, Image as RLImage,
    )

    reg    = ArmarioChaveReserva.query.get_or_404(reg_id)
    arm    = Armario.query.get(reg.armario_id)

    output = BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm,
        topMargin=16*mm, bottomMargin=18*mm,
    )

    W    = A4[0] - 36*mm
    RED  = colors.HexColor("#d40511")
    DARK = colors.HexColor("#111827")
    GRAY = colors.HexColor("#6b7280")
    LGRY = colors.HexColor("#f3f4f6")

    h1  = ParagraphStyle("h1",  fontName="Helvetica-Bold", fontSize=20,
                          textColor=colors.white, alignment=TA_CENTER, leading=26)
    sub = ParagraphStyle("sub", fontName="Helvetica",      fontSize=10,
                          textColor=colors.white, alignment=TA_CENTER, leading=14)
    lbl = ParagraphStyle("lbl", fontName="Helvetica-Bold", fontSize=8,
                          textColor=GRAY, spaceAfter=2, leading=10)
    val = ParagraphStyle("val", fontName="Helvetica-Bold", fontSize=13,
                          textColor=DARK, leading=16)
    sec = ParagraphStyle("sec", fontName="Helvetica-Bold", fontSize=11,
                          textColor=RED, leading=14)
    rod = ParagraphStyle("rod", fontName="Helvetica",      fontSize=8,
                          textColor=GRAY, alignment=TA_CENTER, leading=10)

    def campo(rotulo, texto):
        return [Paragraph(rotulo, lbl), Paragraph(str(texto) if texto else "—", val)]

    def linha_campos(pares, larguras):
        """Recebe lista de (rotulo, valor) e larguras; monta tabela de 2 linhas."""
        row0 = [campo(r, v)[0] for r, v in pares]
        row1 = [campo(r, v)[1] for r, v in pares]
        t = Table([row0, row1], colWidths=larguras)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), LGRY),
            ("TOPPADDING",    (0,0), (-1,-1), 8),
            ("BOTTOMPADDING", (0,0), (-1,-1), 8),
            ("LEFTPADDING",   (0,0), (-1,-1), 10),
            ("RIGHTPADDING",  (0,0), (-1,-1), 10),
        ]))
        return t

    def secao(titulo):
        t = Table([[Paragraph(titulo, sec)]], colWidths=[W])
        t.setStyle(TableStyle([
            ("LINEBELOW",     (0,0), (-1,-1), 1.5, RED),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ]))
        return t

    story = []

    # Cabeçalho vermelho
    cab = Table([[Paragraph("COMPROVANTE DE RETIRADA — CHAVE RESERVA DE ARMÁRIO", h1)]], colWidths=[W])
    cab.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), RED),
        ("TOPPADDING",    (0,0), (-1,-1), 14),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("LEFTPADDING",   (0,0), (-1,-1), 12),
        ("RIGHTPADDING",  (0,0), (-1,-1), 12),
        ("ROUNDEDCORNERS", [8]),
    ]))
    story.append(cab)

    inf = Table([[Paragraph(
        f"DHL SECURITY &nbsp;|&nbsp; {reg.site} &nbsp;|&nbsp; "
        f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", sub,
    )]], colWidths=[W])
    inf.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), DARK),
        ("TOPPADDING",    (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING",   (0,0), (-1,-1), 12),
        ("RIGHTPADDING",  (0,0), (-1,-1), 12),
    ]))
    story.append(inf)
    story.append(Spacer(1, 10*mm))

    # Dados do armário
    story.append(secao("DADOS DO ARMÁRIO"))
    story.append(Spacer(1, 4*mm))
    arm_num   = arm.numero if arm else "?"
    arm_bloco = arm.bloco  if arm else "—"
    arm_colab = arm.colaborador_nome if arm and arm.colaborador_nome else "—"
    story.append(linha_campos(
        [("Nº DO ARMÁRIO", arm_num), ("BLOCO", arm_bloco), ("COLABORADOR ATRIBUÍDO", arm_colab)],
        [W*0.25, W*0.25, W*0.50],
    ))
    story.append(Spacer(1, 2))
    story.append(linha_campos(
        [("DATA E HORA DA RETIRADA", reg.data_retirada.strftime("%d/%m/%Y  %H:%M")),
         ("SITE", reg.site)],
        [W*0.55, W*0.45],
    ))
    story.append(Spacer(1, 8*mm))

    # Dados do retirador
    story.append(secao("DADOS DO RETIRADOR"))
    story.append(Spacer(1, 4*mm))
    story.append(linha_campos(
        [("NOME DO RETIRADOR", reg.retirado_por_nome),
         ("CPF / MATRÍCULA",   reg.retirado_por_cpf)],
        [W*0.55, W*0.45],
    ))
    story.append(Spacer(1, 2))
    story.append(linha_campos(
        [("RESPONSÁVEL PELA ENTREGA (OPERADOR DE SEGURANÇA)", reg.responsavel_entrega)],
        [W],
    ))
    story.append(Spacer(1, 8*mm))

    # Assinatura
    story.append(HRFlowable(width=W, thickness=1, color=colors.HexColor("#e5e7eb")))
    story.append(Spacer(1, 6*mm))
    story.append(secao("ASSINATURA DO RETIRADOR"))
    story.append(Spacer(1, 5*mm))

    if reg.assinatura and reg.assinatura.startswith("data:image"):
        try:
            _, b64data = reg.assinatura.split(",", 1)
            img_buf  = BytesIO(base64.b64decode(b64data))
            sig_img  = RLImage(img_buf, width=120*mm, height=38*mm)
            sig_img.hAlign = "LEFT"
            sig_box = Table([[sig_img]], colWidths=[W])
            sig_box.setStyle(TableStyle([
                ("BACKGROUND",    (0,0), (-1,-1), colors.white),
                ("BOX",           (0,0), (-1,-1), 0.5, colors.HexColor("#d1d5db")),
                ("TOPPADDING",    (0,0), (-1,-1), 6),
                ("BOTTOMPADDING", (0,0), (-1,-1), 6),
                ("ALIGN",         (0,0), (-1,-1), "CENTER"),
            ]))
            story.append(sig_box)
        except Exception:
            story.append(_caixa_sig_vazia(W))
    else:
        story.append(_caixa_sig_vazia(W))

    story.append(Spacer(1, 10*mm))

    # Rodapé
    story.append(HRFlowable(width=W, thickness=0.5, color=colors.HexColor("#e5e7eb")))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        f"Comprovante Nº {reg.id:05d} &nbsp;|&nbsp; DHL Security &nbsp;|&nbsp; "
        f"Documento gerado automaticamente pelo sistema CCTV Control Panel",
        rod,
    ))

    doc.build(story)
    output.seek(0)
    fname = f"comprovante_chave_reserva_{reg.id:05d}.pdf"
    return send_file(output, as_attachment=True, download_name=fname,
                     mimetype="application/pdf")


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORTAR SITUAÇÃO DOS ARMÁRIOS — EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

@armarios_bp.route("/exportar/excel")
@_login_required
def exportar_situacao_excel():
    """Exporta relatório completo dos armários em duas abas:
    Aba 1 — Situação atual de cada armário (LIVRE / OCUPADO + chave reserva)
    Aba 2 — Histórico completo de retiradas de chave reserva
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    site     = session.get("user_site", "")
    is_admin = _is_privileged()

    # Filtros opcionais (query string)
    bloco_f   = request.args.get("bloco",    "").strip()
    status_f  = request.args.get("status",   "").strip().upper()   # LIVRE | OCUPADO | ""
    data_ini  = request.args.get("data_ini", "").strip()
    data_fim  = request.args.get("data_fim", "").strip()

    thin  = Side(style="thin", color="D1D5DB")

    def cabecalho_aba(ws, titulo, subtitulo, n_cols):
        col_letra = chr(ord('A') + n_cols - 1)
        ws.merge_cells(f"A1:{col_letra}1")
        ws["A1"] = titulo
        ws["A1"].font      = Font(size=13, bold=True, color="FFFFFF")
        ws["A1"].fill      = PatternFill("solid", fgColor="D40511")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells(f"A2:{col_letra}2")
        ws["A2"] = subtitulo
        ws["A2"].font      = Font(size=10, bold=True, color="111827")
        ws["A2"].fill      = PatternFill("solid", fgColor="FFCC00")
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 16

    def linha_headers(ws, row, headers):
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.font      = Font(bold=True, size=10, color="111827")
            c.fill      = PatternFill("solid", fgColor="E5E7EB")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws.row_dimensions[row].height = 22

    wb  = Workbook()

    # ── ABA 1: Situação Atual ─────────────────────────────────────────────────
    ws1       = wb.active
    ws1.title = "Situação dos Armários"

    q = Armario.query.filter_by(site=site, ativo=True)
    if bloco_f:
        q = q.filter_by(bloco=bloco_f)
    if status_f in ("LIVRE", "OCUPADO"):
        q = q.filter_by(status=status_f)
    armarios = q.order_by(Armario.bloco, Armario.numero).all()

    total     = len(armarios)
    ocupados  = sum(1 for a in armarios if a.status == "OCUPADO")
    livres    = sum(1 for a in armarios if a.status == "LIVRE")
    chv_fora  = sum(1 for a in armarios if _chave_ativa(a.id))

    filtros_desc = []
    if bloco_f:  filtros_desc.append(f"Bloco: {bloco_f}")
    if status_f: filtros_desc.append(f"Status: {status_f}")
    filtros_txt = " | ".join(filtros_desc) if filtros_desc else "Todos os registros"

    cabecalho_aba(
        ws1,
        f"SITUAÇÃO DOS ARMÁRIOS — {site}",
        (f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
         f" | Usuário: {session.get('user_nome', 'Sistema')}"
         f" | {filtros_txt}"
         f" | Total: {total}  Ocupados: {ocupados}  Livres: {livres}  Chave Fora: {chv_fora}"),
        10,
    )

    headers1 = [
        "Nº Armário", "Bloco", "Status", "Colaborador",
        "CPF / Matrícula", "Atribuído Em", "Atribuído Por",
        "Chave Reserva", "Retirado Por (Chave)", "Data Retirada (Chave)",
    ]
    linha_headers(ws1, 4, headers1)

    fill_ocupado   = PatternFill("solid", fgColor="FEF3C7")
    fill_livre     = PatternFill("solid", fgColor="D1FAE5")
    fill_alerta    = PatternFill("solid", fgColor="FEE2E2")

    for row_i, arm in enumerate(armarios, start=5):
        ch = _chave_ativa(arm.id)
        row_data = [
            arm.numero,
            arm.bloco or "—",
            arm.status,
            arm.colaborador_nome or "—",
            arm.colaborador_cpf  or "—",
            arm.atribuido_em.strftime("%d/%m/%Y %H:%M") if arm.atribuido_em else "—",
            arm.atribuido_por    or "—",
            "RETIRADA" if ch else "NO CLAVICULÁRIO",
            ch.retirado_por_nome if ch else "—",
            ch.data_retirada.strftime("%d/%m/%Y %H:%M") if ch else "—",
        ]
        for col, val in enumerate(row_data, start=1):
            c = ws1.cell(row=row_i, column=col, value=val)
            c.border    = Border(top=thin, left=thin, right=thin, bottom=thin)
            c.alignment = Alignment(vertical="center")

        # Cor da linha
        ch_alerta = ch and (datetime.now() - ch.data_retirada).days >= 1
        if ch_alerta:
            row_fill = fill_alerta
        elif arm.status == "OCUPADO":
            row_fill = fill_ocupado
        else:
            row_fill = fill_livre
        for col in range(1, 11):
            ws1.cell(row=row_i, column=col).fill = row_fill

    for col, w in zip("ABCDEFGHIJ", [14, 14, 12, 28, 18, 20, 24, 16, 28, 22]):
        ws1.column_dimensions[col].width = w
    ws1.freeze_panes = "A5"

    # Legenda
    row_leg = len(armarios) + 7
    ws1.cell(row=row_leg, column=1, value="Legenda:").font = Font(bold=True, size=9)
    for i, (lbl, cor, desc) in enumerate([
        ("LIVRE", "D1FAE5", "Armário disponível"),
        ("OCUPADO", "FEF3C7", "Armário atribuído a colaborador"),
        ("CHAVE FORA (≥1 dia)", "FEE2E2", "Chave reserva retirada há mais de 1 dia — requer atenção"),
    ], start=1):
        ws1.cell(row=row_leg+i, column=1, value="").fill = PatternFill("solid", fgColor=cor)
        ws1.cell(row=row_leg+i, column=1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        ws1.cell(row=row_leg+i, column=2, value=f"{lbl} — {desc}").font = Font(size=9, color="374151")

    # ── ABA 2: Histórico de Chaves Reserva ───────────────────────────────────
    ws2       = wb.create_sheet("Histórico Chave Reserva")

    q2 = ArmarioChaveReserva.query.filter_by(site=site)
    if data_ini:
        try:
            q2 = q2.filter(ArmarioChaveReserva.data_retirada >= datetime.strptime(data_ini, "%Y-%m-%d"))
        except ValueError:
            pass
    if data_fim:
        try:
            dt_f = datetime.strptime(data_fim, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            q2   = q2.filter(ArmarioChaveReserva.data_retirada <= dt_f)
        except ValueError:
            pass
    historico = q2.order_by(ArmarioChaveReserva.data_retirada.desc()).all()

    total_h      = len(historico)
    retiradas_h  = sum(1 for r in historico if r.status == "RETIRADA")
    devolvidas_h = total_h - retiradas_h

    periodo_txt = ""
    if data_ini or data_fim:
        periodo_txt = f" | Período: {data_ini or '—'} a {data_fim or '—'}"

    cabecalho_aba(
        ws2,
        f"HISTÓRICO DE CHAVE RESERVA — {site}",
        (f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
         f" | Usuário: {session.get('user_nome', 'Sistema')}"
         f"{periodo_txt}"
         f" | Total: {total_h}  Em aberto: {retiradas_h}  Devolvidas: {devolvidas_h}"),
        10,
    )

    headers2 = [
        "ID", "Nº Armário", "Bloco", "Retirado Por",
        "CPF / Matrícula", "Resp. Entrega", "Data Retirada",
        "Data Devolução", "Dias em Aberto", "Status",
    ]
    linha_headers(ws2, 4, headers2)

    arm_cache = {}
    agora     = datetime.now()

    fill_ret = PatternFill("solid", fgColor="FEF3C7")
    fill_dev = PatternFill("solid", fgColor="D1FAE5")
    fill_vig = PatternFill("solid", fgColor="FEE2E2")

    for row_i, r in enumerate(historico, start=5):
        if r.armario_id not in arm_cache:
            arm_cache[r.armario_id] = Armario.query.get(r.armario_id)
        arm_r = arm_cache[r.armario_id]
        dias  = ((r.data_devolucao or agora) - r.data_retirada).days

        row_data = [
            r.id,
            arm_r.numero if arm_r else "?",
            arm_r.bloco  if arm_r else "—",
            r.retirado_por_nome,
            r.retirado_por_cpf,
            r.responsavel_entrega,
            r.data_retirada.strftime("%d/%m/%Y %H:%M"),
            r.data_devolucao.strftime("%d/%m/%Y %H:%M") if r.data_devolucao else "—",
            dias,
            r.status,
        ]
        for col, val in enumerate(row_data, start=1):
            c = ws2.cell(row=row_i, column=col, value=val)
            c.border    = Border(top=thin, left=thin, right=thin, bottom=thin)
            c.alignment = Alignment(vertical="center")

        if r.status == "RETIRADA" and dias >= 1:
            row_fill = fill_vig
        elif r.status == "RETIRADA":
            row_fill = fill_ret
        else:
            row_fill = fill_dev
        for col in range(1, 11):
            ws2.cell(row=row_i, column=col).fill = row_fill

    for col, w in zip("ABCDEFGHIJ", [8, 14, 14, 28, 18, 28, 20, 20, 14, 12]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A5"

    # ── Gerar e enviar ────────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    nome_arq = f"armarios_{site}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output, as_attachment=True, download_name=nome_arq,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ═══════════════════════════════════════════════════════════════════════════════
# HISTÓRICO DE EVENTOS — ARMÁRIOS
# ═══════════════════════════════════════════════════════════════════════════════

@armarios_bp.route("/historico")
@_login_required
def historico():
    site     = session.get("user_site", "")
    is_admin = _is_privileged()

    site_f   = request.args.get("site",    "" if is_admin else site)
    arm_f    = request.args.get("armario", "").strip()
    evento_f = request.args.get("evento",  "")
    data_ini = request.args.get("data_ini", "")
    data_fim = request.args.get("data_fim", "")

    q = ArmarioHistorico.query
    if is_admin:
        if site_f:
            q = q.filter_by(site=site_f)
    else:
        q = q.filter_by(site=site)

    if arm_f:
        q = q.filter(ArmarioHistorico.armario_numero.ilike(f"%{arm_f}%"))
    if evento_f:
        q = q.filter_by(evento=evento_f)
    if data_ini:
        try:
            q = q.filter(ArmarioHistorico.data_evento >= datetime.strptime(data_ini, "%Y-%m-%d"))
        except ValueError:
            pass
    if data_fim:
        try:
            dt_f = datetime.strptime(data_fim, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            q = q.filter(ArmarioHistorico.data_evento <= dt_f)
        except ValueError:
            pass

    registros = q.order_by(ArmarioHistorico.data_evento.desc()).limit(500).all()

    sites_lista = []
    if is_admin:
        rows = (ArmarioHistorico.query
                .with_entities(ArmarioHistorico.site)
                .distinct()
                .all())
        sites_lista = sorted({r.site for r in rows if r.site})

    return render_template(
        "armarios/historico.html",
        registros=registros,
        site=site,
        is_admin=is_admin,
        sites_lista=sites_lista,
        site_f=site_f,
        arm_f=arm_f,
        evento_f=evento_f,
        data_ini=data_ini,
        data_fim=data_fim,
    )


@armarios_bp.route("/historico/exportar")
@_login_required
def exportar_historico():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    site     = session.get("user_site", "")
    is_admin = _is_privileged()

    site_f   = request.args.get("site",    "" if is_admin else site)
    arm_f    = request.args.get("armario", "").strip()
    evento_f = request.args.get("evento",  "")
    data_ini = request.args.get("data_ini", "")
    data_fim = request.args.get("data_fim", "")

    q = ArmarioHistorico.query
    if is_admin:
        if site_f:
            q = q.filter_by(site=site_f)
    else:
        q = q.filter_by(site=site)

    if arm_f:
        q = q.filter(ArmarioHistorico.armario_numero.ilike(f"%{arm_f}%"))
    if evento_f:
        q = q.filter_by(evento=evento_f)
    if data_ini:
        try:
            q = q.filter(ArmarioHistorico.data_evento >= datetime.strptime(data_ini, "%Y-%m-%d"))
        except ValueError:
            pass
    if data_fim:
        try:
            dt_f = datetime.strptime(data_fim, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            q = q.filter(ArmarioHistorico.data_evento <= dt_f)
        except ValueError:
            pass

    registros = q.order_by(ArmarioHistorico.data_evento.desc()).all()

    thin = Side(style="thin", color="D1D5DB")
    wb   = Workbook()
    ws   = wb.active
    ws.title = "Histórico de Armários"

    n_cols = 9
    col_letra = chr(ord('A') + n_cols - 1)

    filtros = []
    if site_f:   filtros.append(f"Site: {site_f}")
    if arm_f:    filtros.append(f"Armário: {arm_f}")
    if evento_f: filtros.append(f"Evento: {evento_f}")
    if data_ini: filtros.append(f"De: {data_ini}")
    if data_fim: filtros.append(f"Até: {data_fim}")
    filtros_txt = " | ".join(filtros) if filtros else "Todos os registros"

    ws.merge_cells(f"A1:{col_letra}1")
    ws["A1"] = f"HISTÓRICO DE ARMÁRIOS — {(site_f or site).upper()}"
    ws["A1"].font      = Font(size=13, bold=True, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="D40511")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{col_letra}2")
    ws["A2"] = (f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                f" | Usuário: {session.get('user_nome', 'Sistema')}"
                f" | {filtros_txt}"
                f" | Total: {len(registros)} registro(s)")
    ws["A2"].font      = Font(size=10, bold=True, color="111827")
    ws["A2"].fill      = PatternFill("solid", fgColor="FFCC00")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    headers = ["Data/Hora", "Site", "Nº Armário", "Bloco",
               "Evento", "Colaborador", "CPF / Matrícula", "Operador", "Observação"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font      = Font(bold=True, size=10, color="111827")
        c.fill      = PatternFill("solid", fgColor="E5E7EB")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws.row_dimensions[4].height = 22

    cor_evento = {
        'ATRIBUIÇÃO':    "D1FAE5",
        'LIBERAÇÃO':     "FEE2E2",
        'CHAVE RETIRADA': "FEF3C7",
        'CHAVE DEVOLVIDA': "EDE9FE",
    }

    for row_i, r in enumerate(registros, start=5):
        row_data = [
            r.data_evento.strftime("%d/%m/%Y %H:%M") if r.data_evento else "—",
            r.site or "—",
            r.armario_numero or "—",
            r.bloco or "—",
            r.evento or "—",
            r.colaborador_nome or "—",
            r.colaborador_cpf  or "—",
            r.operador         or "—",
            r.observacao       or "—",
        ]
        fill_hex = cor_evento.get(r.evento, "FFFFFF")
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_i, column=col, value=val)
            c.border    = Border(top=thin, left=thin, right=thin, bottom=thin)
            c.alignment = Alignment(vertical="center")
            c.fill      = PatternFill("solid", fgColor=fill_hex)

    for col, w in zip("ABCDEFGHI", [18, 14, 14, 12, 18, 28, 18, 22, 30]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    nome_arq = f"ARM-HIST-{(site_f or site).replace(' ','_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output, as_attachment=True, download_name=nome_arq,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _caixa_sig_vazia(W):
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    box = Table([[""]], colWidths=[W], rowHeights=[42*mm])
    box.setStyle(TableStyle([
        ("BOX",        (0,0), (-1,-1), 0.8, colors.HexColor("#cbd5e1")),
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#f8fafc")),
        ("LINEBELOW",  (0,0), (-1,-1), 1.5, colors.HexColor("#94a3b8")),
    ]))
    return box
