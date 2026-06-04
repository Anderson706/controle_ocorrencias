# achados_blueprint.py — Achados e Perdidos integrado ao CCTV Control Panel
# Registrado em app.py com url_prefix='/achados'
# Autenticação usa a sessão do CCTV (session["user_id"])

from flask import (
    Blueprint, render_template, request, redirect,
    url_for, flash, session, send_file, current_app
)
from functools import wraps
from datetime import datetime, timedelta
from io import BytesIO
import os, subprocess, tempfile

from werkzeug.utils import secure_filename

from reportlab.lib.pagesizes import A4
from reportlab.lib.colors import HexColor
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.lib.utils import ImageReader

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment as XLAlign

achados_bp = Blueprint('achados', __name__, template_folder='templates')

# ── Globals preenchidos via setup_achados ─────────────────────────────────────
_db           = None
AchadoPerdido = None

ALLOWED_EXT = {"png", "jpg", "jpeg", "webp", "gif"}
_STATUS_VALIDOS = ("Pendente", "Entregue", "Devolvido", "Doacao")


# ── Inicialização ──────────────────────────────────────────────────────────────
def setup_achados(db):
    global _db, AchadoPerdido
    _db = db

    class _AchadoPerdido(db.Model):
        __tablename__ = "ACHADOS_PERDIDOS"

        id           = db.Column(db.Integer, db.Identity(start=1), primary_key=True)
        id_registro  = db.Column(db.String(50),  nullable=False)   # ex: SHEINGUA-2026-0001
        numero_site  = db.Column(db.Integer,     nullable=True)     # sequência por site/ano
        id_anterior  = db.Column(db.String(50),  nullable=True)
        objeto      = db.Column(db.String(200),  nullable=False)
        responsavel = db.Column(db.String(150),  nullable=False)
        data        = db.Column(db.String(20),   nullable=False)   # YYYY-MM-DD
        turno       = db.Column(db.String(30),    nullable=False)
        descricao   = db.Column(db.Text,         nullable=True)
        foto_path   = db.Column(db.String(500),  nullable=True)
        status      = db.Column(db.String(30),   nullable=False, default='Pendente')
        retirado_por = db.Column(db.String(150), nullable=True)
        site        = db.Column(db.String(128),  nullable=True)
        criado_por  = db.Column(db.String(120),  nullable=True)
        created_at  = db.Column(db.DateTime,     nullable=False, default=datetime.utcnow)

    AchadoPerdido = _AchadoPerdido


# ── Auth decorator ─────────────────────────────────────────────────────────────
def _login_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            flash("Faça login para acessar.", "danger")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapped


# ── Helpers de site ────────────────────────────────────────────────────────────
def _sites_usuario():
    """Retorna lista de sites permitidos para o usuário logado."""
    perfil = (session.get("user_perfil") or "").upper()
    if perfil == "ADMIN":
        return []
    uid = session.get("user_id")
    if uid:
        from sqlalchemy import text as _text
        rows = _db.session.execute(
            _text("SELECT SITE_NOME FROM USUARIO_SITES WHERE USUARIO_ID = :uid"),
            {"uid": uid}
        ).fetchall()
        if rows:
            return [r[0] for r in rows]
    site = session.get("user_site", "")
    return [site] if site else []


def _filtrar_query(query):
    """Aplica filtro de site(s) na query."""
    sites = _sites_usuario()
    if not sites:
        return query
    if len(sites) == 1:
        return query.filter(AchadoPerdido.site == sites[0])
    return query.filter(AchadoPerdido.site.in_(sites))


# ── Upload ─────────────────────────────────────────────────────────────────────
def _upload_folder():
    folder = os.path.join(current_app.static_folder, "uploads", "achados_perdidos")
    os.makedirs(folder, exist_ok=True)
    return folder


def _allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXT


# ── Outlook ────────────────────────────────────────────────────────────────────
def _abrir_outlook(caminho_arquivo: str, assunto: str, corpo_html: str):
    ass = assunto.replace("'", "''")
    bod = corpo_html.replace("'", "''").replace("\n", " ")
    arq = caminho_arquivo.replace("\\", "\\\\")
    ps = (
        f"$ol = New-Object -ComObject Outlook.Application; "
        f"$m  = $ol.CreateItem(0); "
        f"$m.Subject  = '{ass}'; "
        f"$m.HTMLBody = '{bod}'; "
        f"$m.Attachments.Add('{arq}'); "
        f"$m.Display()"
    )
    subprocess.Popen(
        ["powershell", "-ExecutionPolicy", "Bypass", "-Command", ps],
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000),
    )


# ── Gerador de PDF ─────────────────────────────────────────────────────────────
def _gerar_pdf(r: dict) -> BytesIO:
    buf = BytesIO()
    largura, altura = A4
    c = rl_canvas.Canvas(buf, pagesize=A4)

    DHL_YELLOW     = HexColor("#ffcc00")
    DHL_RED        = HexColor("#d40511")
    DHL_BLACK      = HexColor("#000000")
    DHL_GRAY_LIGHT = HexColor("#f4f4f4")
    DHL_BORDER     = HexColor("#dddddd")

    c.setTitle(f"Achados e Perdidos - Registro #{r['id_registro']}")
    c.setAuthor("Security & Loss Prevention - DHL")

    def wrap_text(text, font_name, font_size, max_width):
        palavras = (text or "").split()
        linhas, linha_atual = [], ""
        for p in palavras:
            teste = (linha_atual + " " + p).strip()
            if stringWidth(teste, font_name, font_size) <= max_width:
                linha_atual = teste
            else:
                if linha_atual:
                    linhas.append(linha_atual)
                linha_atual = p
        if linha_atual:
            linhas.append(linha_atual)
        return linhas or [""]

    # Barra lateral vermelha
    c.setFillColor(DHL_RED)
    c.rect(0, 0, 8, altura, fill=True, stroke=False)

    # Marca d'água
    c.saveState()
    c.setFont("Helvetica-Bold", 40)
    c.setFillColor(HexColor("#f3f3f3"))
    c.translate(largura / 2, altura / 2)
    c.rotate(35)
    c.drawCentredString(0, 0, "DHL SECURITY")
    c.restoreState()

    # Cabeçalho amarelo
    topo = 70
    c.setFillColor(DHL_YELLOW)
    c.rect(0, altura - topo, largura, topo, fill=True, stroke=False)
    c.setFillColor(DHL_RED)
    c.rect(0, altura - topo - 4, largura, 4, fill=True, stroke=False)

    logo_path = os.path.join(current_app.static_folder, "logo.png")
    if os.path.exists(logo_path):
        try:
            c.drawImage(logo_path, 30, altura - topo + 10,
                        width=120, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

    c.setFont("Helvetica-Bold", 16)
    c.setFillColor(DHL_BLACK)
    c.drawCentredString(largura / 2, altura - topo + 25, "Registro de Achados e Perdidos")

    c.setFont("Helvetica-Bold", 12)
    c.drawRightString(largura - 40, altura - topo + 20, f"Registro nº {r['id_registro']}")

    c.setStrokeColor(DHL_RED)
    c.setLineWidth(1.5)
    c.line(40, altura - topo - 15, largura - 40, altura - topo - 15)

    margem_esq = 50
    margem_dir = largura - 50
    esp = 14
    y = altura - topo - 60

    # Card: Dados do registro
    turno_label = {"1": "Turno 1 (06h–14h)", "2": "Turno 2 (14h–22h)", "3": "Turno 3 (22h–06h)"}
    dados = [
        ("Objeto",        r.get("objeto", "")),
        ("Responsável",   r.get("responsavel", "")),
        ("Data",          r.get("data", "")),
        ("Turno",         turno_label.get(str(r.get("turno", "")), str(r.get("turno", "")))),
        ("Site",          r.get("site") or "—"),
        ("Status",        r.get("status") or "—"),
        ("Retirado por",  r.get("retirado_por") or "—"),
    ]
    pad = 14
    card_x = margem_esq - 10
    card_w = margem_dir - margem_esq + 20
    altura_card = pad * 2 + (len(dados) + 1) * esp
    card_y = y - altura_card + pad

    c.setStrokeColor(DHL_BORDER); c.setFillColor(DHL_GRAY_LIGHT)
    c.roundRect(card_x, card_y, card_w, altura_card, 10, stroke=True, fill=True)

    c.setFont("Helvetica-Bold", 12); c.setFillColor(DHL_RED)
    c.drawString(margem_esq, y - esp, "Dados do Registro")
    linha_y = y - esp * 2
    for lbl, val in dados:
        c.setFont("Helvetica-Bold", 11); c.setFillColor(DHL_BLACK)
        c.drawString(margem_esq, linha_y, f"{lbl}:")
        c.setFont("Helvetica", 11)
        c.drawString(margem_esq + 95, linha_y, str(val) if val is not None else "—")
        linha_y -= esp
    y = card_y - 30

    # Card: Descrição
    desc_linhas = wrap_text(r.get("descricao", ""), "Helvetica", 11, margem_dir - margem_esq)
    alt_desc = pad * 2 + (len(desc_linhas) + 2) * esp
    card_desc_y = y - alt_desc + pad
    c.setStrokeColor(DHL_BORDER); c.setFillColor(HexColor("#ffffff"))
    c.roundRect(card_x, card_desc_y, card_w, alt_desc, 10, stroke=True, fill=True)
    c.setFont("Helvetica-Bold", 12); c.setFillColor(DHL_RED)
    c.drawString(margem_esq, y - esp, "Descrição detalhada")
    c.setFont("Helvetica", 11); c.setFillColor(DHL_BLACK)
    l_y = y - esp * 2
    for ln in desc_linhas:
        c.drawString(margem_esq, l_y, ln)
        l_y -= esp
    y = card_desc_y - 30

    # Card: Foto
    foto_rel = r.get("foto_path")
    if foto_rel:
        foto_abs = os.path.join(current_app.static_folder, foto_rel)
        if os.path.exists(foto_abs):
            try:
                img = ImageReader(foto_abs)
                iw, ih = img.getSize()
                max_w, max_h = 260, 180
                esc = min(max_w / iw, max_h / ih)
                dw, dh = iw * esc, ih * esc
                alt_img = pad * 2 + 2 * esp + dh + 10
                card_img_y = y - alt_img + pad
                c.setStrokeColor(DHL_BORDER); c.setFillColor(DHL_GRAY_LIGHT)
                c.roundRect(card_x, card_img_y, card_w, alt_img, 10, stroke=True, fill=True)
                c.setFont("Helvetica-Bold", 12); c.setFillColor(DHL_RED)
                c.drawString(margem_esq, y - esp, "Imagem do objeto")
                xi = (largura - dw) / 2
                yi = card_img_y + pad + esp
                c.setStrokeColor(DHL_BLACK)
                c.rect(xi - 3, yi - 3, dw + 6, dh + 6, stroke=True, fill=False)
                c.drawImage(img, xi, yi, width=dw, height=dh,
                            preserveAspectRatio=True, mask='auto')
                c.setFont("Helvetica-Oblique", 9); c.setFillColor(DHL_BLACK)
                c.drawCentredString(largura / 2, card_img_y + pad, "Foto do item registrado.")
                y = card_img_y - 25
            except Exception:
                pass

    # Assinaturas
    c.setStrokeColor(DHL_BORDER); c.setLineWidth(0.8)
    c.line(margem_esq, 110, 260, 110)
    c.setFont("Helvetica", 8); c.setFillColor(DHL_BLACK)
    c.drawString(margem_esq, 114, "Responsável Security / Data / Assinatura")
    c.line(320, 110, largura - 50, 110)
    c.drawString(320, 114, "Responsável Operação / Data / Assinatura")

    # Rodapé amarelo
    c.setFillColor(DHL_YELLOW)
    c.rect(0, 0, largura, 32, fill=True, stroke=False)
    c.setFont("Helvetica-Bold", 9); c.setFillColor(DHL_BLACK)
    c.drawString(40, 16, "Security & Loss Prevention - DHL")
    c.setFont("Helvetica", 8)
    c.drawRightString(largura - 40, 18,
                      f"Documento gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.setFont("Helvetica-Oblique", 7)
    c.drawCentredString(largura / 2, 6,
                        "Documento de uso interno – Proibida a reprodução não autorizada.")

    c.showPage()
    c.save()
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════════════
# ROTAS
# ═══════════════════════════════════════════════════════════════════════════════

def _gerar_codigo_achado(site: str):
    """Gera o próximo id_registro no formato SITECODE-ANO-SEQ (ex: SHEINGUA-2026-0001).
    Retorna (codigo_str, seq_int).
    """
    import re as _re
    from sqlalchemy import func as _func
    clean = _re.sub(r'[^A-Z0-9]', '', (site or "SITE").upper())[:8] or "SITE"
    ano   = datetime.now().year
    max_seq = (
        _db.session.query(_func.max(AchadoPerdido.numero_site))
        .filter(AchadoPerdido.site == site,
                _db.func.substr(AchadoPerdido.id_registro, 1, 4+len(clean)+1)
                == f"{clean}-{ano}")
        .scalar()
        or 0
    )
    seq = max_seq + 1
    return f"{clean}-{ano}-{seq:04d}", seq


@achados_bp.route("/novo", methods=["GET"])
@_login_required
def novo():
    site       = session.get("user_site") or ""
    next_id, _ = _gerar_codigo_achado(site)
    today      = datetime.now().strftime("%Y-%m-%d")
    return render_template("achados_novo.html", next_id=next_id, today=today)


@achados_bp.route("/salvar", methods=["POST"])
@_login_required
def salvar():
    objeto      = (request.form.get("objeto") or "").strip()
    data        = (request.form.get("data") or "").strip()
    turno       = (request.form.get("turno") or "").strip()
    descricao   = (request.form.get("descricao") or "").strip()

    if not all([objeto, data, turno, descricao]):
        flash("Preencha todos os campos obrigatórios.", "danger")
        return redirect(url_for("achados.novo"))

    # Campos gerados no servidor — nunca confia no formulário
    site                    = session.get("user_site") or ""
    responsavel             = session.get("user_nome") or ""
    id_registro, num_site   = _gerar_codigo_achado(site)

    foto_path = None
    foto = request.files.get("foto")
    if foto and foto.filename:
        if not _allowed_file(foto.filename):
            flash("Formato não permitido. Use PNG / JPG / JPEG / WEBP / GIF.", "danger")
            return redirect(url_for("achados.novo"))
        safe  = secure_filename(foto.filename)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"AP_{id_registro}_{stamp}_{safe}"
        foto.save(os.path.join(_upload_folder(), fname))
        foto_path = f"uploads/achados_perdidos/{fname}"

    reg = AchadoPerdido(
        id_registro  = id_registro,
        numero_site  = num_site,
        objeto       = objeto,
        responsavel  = responsavel,
        data         = data,
        turno        = turno,
        descricao    = descricao,
        foto_path    = foto_path,
        status       = "Pendente",
        site         = site,
        criado_por   = session.get("user_nome") or "",
    )
    _db.session.add(reg)
    _db.session.commit()

    flash(f"Registro #{id_registro} — {objeto} — salvo com sucesso!", "success")
    return redirect(url_for("achados.lista"))


@achados_bp.route("/lista", methods=["GET"])
@_login_required
def lista():
    q     = (request.args.get("q") or "").strip()
    f_status = (request.args.get("status") or "").strip()

    query = AchadoPerdido.query
    query = _filtrar_query(query)

    if q:
        like = f"%{q}%"
        from sqlalchemy import or_
        query = query.filter(or_(
            AchadoPerdido.objeto.ilike(like),
            AchadoPerdido.responsavel.ilike(like),
            AchadoPerdido.descricao.ilike(like),
        ))

    if f_status:
        query = query.filter(AchadoPerdido.status == f_status)

    registros_raw = query.order_by(AchadoPerdido.created_at.desc()).all()

    registros = []
    hoje = datetime.now().date()
    for r in registros_raw:
        d = r.__dict__.copy()
        try:
            dt = datetime.strptime(r.data, "%Y-%m-%d").date()
            prazo = dt + timedelta(days=90)
            d["prazo_entrega"] = prazo.isoformat()
            d["prazo_vencido"] = prazo < hoje
        except Exception:
            d["prazo_entrega"] = ""
            d["prazo_vencido"] = False
        registros.append(d)

    total_pendente = sum(1 for r in registros_raw if r.status == "Pendente")

    return render_template(
        "achados_lista.html",
        registros=registros,
        q=q,
        f_status=f_status,
        total_pendente=total_pendente,
    )


@achados_bp.route("/<int:row_id>/status", methods=["POST"])
@_login_required
def atualizar_status(row_id):
    reg = AchadoPerdido.query.get_or_404(row_id)
    novo_status  = (request.form.get("status") or "").strip()
    retirado_por = (request.form.get("retirado_por") or "").strip() or None

    if novo_status not in _STATUS_VALIDOS:
        flash("Status inválido.", "danger")
        return redirect(url_for("achados.lista"))

    reg.status      = novo_status
    reg.retirado_por = retirado_por
    _db.session.commit()
    flash("Status atualizado com sucesso!", "success")
    return redirect(url_for("achados.lista"))


@achados_bp.route("/<int:row_id>/editar", methods=["GET"])
@_login_required
def editar(row_id):
    reg = AchadoPerdido.query.get_or_404(row_id)
    return render_template("achados_editar.html", reg=reg)


@achados_bp.route("/<int:row_id>/editar/salvar", methods=["POST"])
@_login_required
def editar_salvar(row_id):
    reg = AchadoPerdido.query.get_or_404(row_id)
    reg.objeto      = (request.form.get("objeto") or "").strip()
    reg.responsavel = (request.form.get("responsavel") or "").strip()
    reg.descricao   = (request.form.get("descricao") or "").strip()
    reg.turno       = (request.form.get("turno") or reg.turno or "").strip()
    reg.status      = (request.form.get("status") or reg.status).strip()

    # Troca de foto (opcional)
    nova_foto = request.files.get("foto")
    if nova_foto and nova_foto.filename:
        if _allowed_file(nova_foto.filename):
            safe  = secure_filename(nova_foto.filename)
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            fname = f"AP_{reg.id_registro}_{stamp}_{safe}"
            nova_foto.save(os.path.join(_upload_folder(), fname))
            reg.foto_path = f"uploads/achados_perdidos/{fname}"

    _db.session.commit()
    flash("Registro atualizado com sucesso!", "success")
    return redirect(url_for("achados.lista"))


@achados_bp.route("/<int:row_id>/excluir", methods=["POST"])
@_login_required
def excluir(row_id):
    reg = AchadoPerdido.query.get_or_404(row_id)
    _db.session.delete(reg)
    _db.session.commit()
    flash("Registro excluído.", "danger")
    return redirect(url_for("achados.lista"))


@achados_bp.route("/<int:row_id>/pdf")
@_login_required
def pdf(row_id):
    reg = AchadoPerdido.query.get_or_404(row_id)
    r   = {c.name: getattr(reg, c.name) for c in reg.__table__.columns}
    buf = _gerar_pdf(r)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"AP_{reg.id_registro}_{reg.objeto[:30]}.pdf".replace(" ", "_"),
        mimetype="application/pdf",
    )


@achados_bp.route("/<int:row_id>/email")
@_login_required
def enviar_email(row_id):
    reg = AchadoPerdido.query.get_or_404(row_id)
    r   = {c.name: getattr(reg, c.name) for c in reg.__table__.columns}

    # Gera PDF temporário
    buf = _gerar_pdf(r)
    nome_arq = f"AP_{reg.id_registro} - {reg.objeto[:40]}.pdf".replace(" ", "_")
    tmp_path = os.path.join(tempfile.gettempdir(), nome_arq)
    with open(tmp_path, "wb") as f:
        f.write(buf.read())

    # Data formatada
    try:
        data_fmt = datetime.strptime(reg.data, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        data_fmt = reg.data or "—"

    turno_label = {"1": "Turno 1 (06h–14h)", "2": "Turno 2 (14h–22h)", "3": "Turno 3 (22h–06h)"}
    assunto = f"[Achados e Perdidos] Registro #{reg.id_registro} - {reg.objeto}"

    corpo = (
        '<div style="font-family:Arial,sans-serif;background:#f4f4f4;padding:20px;">'
        '<div style="max-width:520px;margin:0 auto;background:#fff;border-radius:8px;'
        'overflow:hidden;box-shadow:0 4px 10px rgba(0,0,0,.10);">'

        '<div style="background:#FFCC00;border-bottom:4px solid #D40511;padding:20px 24px;">'
        '<div style="font-size:11px;font-weight:900;letter-spacing:1px;color:#111;margin-bottom:4px;">DHL SECURITY</div>'
        '<div style="font-size:20px;font-weight:900;color:#1A1A1A;">Achados e Perdidos</div>'
        f'<div style="font-size:13px;color:#333;margin-top:2px;">Registro #{reg.id_registro}</div>'
        '</div>'

        '<div style="padding:24px;">'
        '<p style="color:#374151;font-size:14px;margin-top:0;">Prezados,</p>'
        '<p style="color:#374151;font-size:14px;">Segue em anexo o registro de '
        '<strong>Achados e Perdidos</strong>. Confira os dados abaixo:</p>'

        '<table style="width:100%;border-collapse:collapse;background:#f8fafc;'
        'border:1px solid #e5e7eb;border-radius:8px;font-size:13px;margin:16px 0;">'
        f'<tr><td style="padding:10px 14px;border-bottom:1px solid #e5e7eb;color:#6b7280;font-weight:700;width:40%;">Nº Registro</td>'
        f'<td style="padding:10px 14px;border-bottom:1px solid #e5e7eb;color:#1f2937;font-weight:800;">#{reg.id_registro}</td></tr>'
        f'<tr><td style="padding:10px 14px;border-bottom:1px solid #e5e7eb;color:#6b7280;font-weight:700;">Objeto</td>'
        f'<td style="padding:10px 14px;border-bottom:1px solid #e5e7eb;color:#1f2937;font-weight:800;">{reg.objeto}</td></tr>'
        f'<tr><td style="padding:10px 14px;border-bottom:1px solid #e5e7eb;color:#6b7280;font-weight:700;">Responsável</td>'
        f'<td style="padding:10px 14px;border-bottom:1px solid #e5e7eb;color:#1f2937;font-weight:800;">{reg.responsavel}</td></tr>'
        f'<tr><td style="padding:10px 14px;border-bottom:1px solid #e5e7eb;color:#6b7280;font-weight:700;">Data</td>'
        f'<td style="padding:10px 14px;border-bottom:1px solid #e5e7eb;color:#1f2937;font-weight:800;">{data_fmt}</td></tr>'
        f'<tr><td style="padding:10px 14px;border-bottom:1px solid #e5e7eb;color:#6b7280;font-weight:700;">Turno</td>'
        f'<td style="padding:10px 14px;border-bottom:1px solid #e5e7eb;color:#1f2937;font-weight:800;">{turno_label.get(str(reg.turno) if reg.turno is not None else "", str(reg.turno or ""))}</td></tr>'
        f'<tr><td style="padding:10px 14px;color:#6b7280;font-weight:700;">Status</td>'
        f'<td style="padding:10px 14px;color:#1f2937;font-weight:800;">{reg.status}</td></tr>'
        '</table>'

        '<p style="color:#374151;font-size:14px;">Colocamo-nos à disposição para esclarecimentos.</p>'
        '</div>'

        '<div style="background:#f0f0f0;text-align:center;padding:12px;font-size:11px;color:#9ca3af;">'
        'DHL Supply Chain · Security &amp; Loss Prevention · CCTV Control Panel · Uso interno'
        '</div>'
        '</div></div>'
    )

    _abrir_outlook(tmp_path, assunto, corpo)
    flash("Outlook aberto com o registro anexado. Preencha os destinatários e envie.", "success")
    return redirect(url_for("achados.lista"))


@achados_bp.route("/dashboard")
@_login_required
def dashboard():
    from sqlalchemy import func as _func, case as _case

    base = AchadoPerdido.query
    base = _filtrar_query(base)
    todos = base.all()

    total      = len(todos)
    pendentes  = sum(1 for r in todos if r.status == "Pendente")
    resolvidos = sum(1 for r in todos if r.status in ("Entregue", "Devolvido"))
    doacoes    = sum(1 for r in todos if r.status == "Doacao")

    # Prazo vencido (pendentes com mais de 90 dias)
    hoje = datetime.now().date()
    vencidos = 0
    for r in todos:
        if r.status == "Pendente":
            try:
                dt = datetime.strptime(r.data, "%Y-%m-%d").date()
                if dt + timedelta(days=90) < hoje:
                    vencidos += 1
            except Exception:
                pass

    # Ranking por responsável
    from collections import Counter
    cnt = Counter(r.responsavel for r in todos)
    ranking = [{"responsavel": k, "qtd": v} for k, v in cnt.most_common(6)]
    top = ranking[0] if ranking else None

    # Distribuição por status
    status_labels = ["Pendente", "Entregue", "Devolvido", "Doacao"]
    status_counts = [sum(1 for r in todos if r.status == s) for s in status_labels]

    # Por mês (últimos 6 meses)
    from collections import defaultdict
    por_mes = defaultdict(int)
    for r in todos:
        try:
            mes = datetime.strptime(r.data, "%Y-%m-%d").strftime("%m/%Y")
            por_mes[mes] += 1
        except Exception:
            pass
    meses_sorted = sorted(por_mes.keys(),
                          key=lambda m: datetime.strptime(m, "%m/%Y"))[-6:]
    meses_labels = meses_sorted
    meses_data   = [por_mes[m] for m in meses_sorted]

    return render_template(
        "achados_dashboard.html",
        total=total,
        pendentes=pendentes,
        resolvidos=resolvidos,
        doacoes=doacoes,
        vencidos=vencidos,
        ranking=ranking,
        top=top,
        status_labels=status_labels,
        status_counts=status_counts,
        meses_labels=meses_labels,
        meses_data=meses_data,
    )


@achados_bp.route("/exportar/excel")
@_login_required
def exportar_excel():
    query = AchadoPerdido.query
    query = _filtrar_query(query)
    rows  = query.order_by(AchadoPerdido.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Achados e Perdidos"

    headers = ["ID Reg.", "Objeto", "Responsável", "Data", "Turno",
               "Descrição", "Status", "Retirado Por", "Site", "Criado Por", "Criado Em"]

    # Cabeçalho com estilo DHL
    hdr_fill = PatternFill("solid", fgColor="FFCC00")
    hdr_font = Font(bold=True, color="000000")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = XLAlign(horizontal="center")

    turno_label = {"1": "Turno 1", "2": "Turno 2", "3": "Turno 3"}
    for r in rows:
        ws.append([
            r.id_registro, r.objeto, r.responsavel, r.data,
            turno_label.get(str(r.turno) if r.turno is not None else "", str(r.turno or "")),
            r.descricao, r.status, r.retirado_por or "",
            r.site or "", r.criado_por or "",
            r.created_at.strftime("%d/%m/%Y %H:%M") if r.created_at else "",
        ])

    # Ajusta largura das colunas
    col_widths = [10, 28, 22, 12, 14, 45, 12, 22, 18, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = datetime.now().strftime("achados_perdidos_%Y%m%d_%H%M%S.xlsx")
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
