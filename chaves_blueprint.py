# chaves_blueprint.py — Controle de Chaves integrado ao CCTV Control Panel
# Registrado em app.py com url_prefix='/chaves'
# Autenticação usa a sessão do CCTV (session["user_id"])
# Páginas: Meu Claviculário · Realizar Retirada · Realizar Devolução

from flask import (
    Blueprint, render_template, request, redirect,
    url_for, flash, session, send_file, jsonify
)
from functools import wraps
from datetime import datetime, timedelta
from io import BytesIO
import base64

chaves_bp = Blueprint('chaves', __name__, template_folder='templates')


def _tempo_decorrido(dt):
    """Retorna string legível do tempo decorrido desde dt.
    Ex: '5d 3h 20min', '2h 45min', '38min'
    """
    delta = datetime.now() - dt
    total_seg = int(delta.total_seconds())
    if total_seg < 0:
        return "agora"
    dias  = total_seg // 86400
    horas = (total_seg % 86400) // 3600
    mins  = (total_seg % 3600) // 60
    partes = []
    if dias:
        partes.append(f"{dias}d")
    if horas:
        partes.append(f"{horas}h")
    if mins or not partes:
        partes.append(f"{mins}min")
    return " ".join(partes)

# ── Referências globais (preenchidas via setup_chaves) ─────────────────────────
_db              = None
ClavicularioChave    = None
ClavicularioRetirada = None


# ── Inicialização (chamada de app.py após db = SQLAlchemy(app)) ────────────────
def setup_chaves(db):
    global _db, ClavicularioChave, ClavicularioRetirada
    _db = db

    # ── Models ────────────────────────────────────────────────────────────────

    class _ClavicularioChave(db.Model):
        """Cadastro de chaves físicas de um site (gerenciado por ADMIN)."""
        __tablename__ = "claviculario_chave"

        id           = db.Column(db.Integer,
                           db.Sequence('clav_chave_id_seq', start=1),
                           primary_key=True)
        numero_chave = db.Column(db.String(50),  nullable=False)
        local        = db.Column(db.String(150), nullable=False)
        site         = db.Column(db.String(100), nullable=False)
        criador_id   = db.Column(db.Integer,     nullable=False)
        criador_nome = db.Column(db.String(150), nullable=False)
        criado_em    = db.Column(db.DateTime, default=datetime.now)
        ativa        = db.Column(db.Boolean,  default=True)

    class _ClavicularioRetirada(db.Model):
        """Registro de retirada/devolução de uma chave do claviculário."""
        __tablename__ = "claviculario_retirada"

        id                  = db.Column(db.Integer,
                                db.Sequence('clav_retirada_id_seq', start=1),
                                primary_key=True)
        chave_id            = db.Column(db.Integer,     db.ForeignKey("claviculario_chave.id"), nullable=False)
        # Retirador — quem está pegando a chave (preenchido manualmente)
        cpf_matricula       = db.Column(db.String(50),  nullable=False)
        nome_retirador      = db.Column(db.String(150), nullable=False, default="")
        # Responsável pela entrega — usuário logado (automático)
        responsavel_entrega = db.Column(db.String(150), nullable=False)
        usuario_id          = db.Column(db.Integer,     nullable=False)
        usuario_nome        = db.Column(db.String(150), nullable=False, default="")
        # Dados gerais
        site                = db.Column(db.String(100), nullable=False)
        data_retirada       = db.Column(db.DateTime, nullable=False, default=datetime.now)
        data_devolucao      = db.Column(db.DateTime, nullable=True)
        status              = db.Column(db.String(30),  nullable=False, default="EM USO")
        # Assinatura do retirador (base64 do canvas PNG)
        assinatura          = db.Column(db.Text, nullable=True)

    ClavicularioChave    = _ClavicularioChave
    ClavicularioRetirada = _ClavicularioRetirada

    import sys
    _mod = sys.modules[__name__]
    _mod.ClavicularioChave    = _ClavicularioChave
    _mod.ClavicularioRetirada = _ClavicularioRetirada
    _mod._db                  = db


# ── Decorators ────────────────────────────────────────────────────────────────

def login_required_chaves(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            flash("Faça login para acessar o sistema.", "danger")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if (session.get("user_perfil") or "").upper() != "ADMIN":
            flash("Acesso restrito a administradores.", "danger")
            return redirect(url_for("chaves.meu_claviculario"))
        return f(*args, **kwargs)
    return decorated


# ── Entrada do módulo ─────────────────────────────────────────────────────────

@chaves_bp.route("/")
@login_required_chaves
def index():
    return redirect(url_for("chaves.meu_claviculario"))


# ═══════════════════════════════════════════════════════════════════════════════
# MEU CLAVICULÁRIO — Cadastro de chaves do site
# ═══════════════════════════════════════════════════════════════════════════════

@chaves_bp.route("/meu-claviculario", methods=["GET", "POST"])
@login_required_chaves
def meu_claviculario():
    site     = session.get("user_site", "")
    is_admin = (session.get("user_perfil") or "").upper() == "ADMIN"

    if request.method == "POST":
        if not is_admin:
            flash("Apenas administradores podem adicionar chaves ao claviculário.", "danger")
            return redirect(url_for("chaves.meu_claviculario"))

        numero_chave = request.form.get("numero_chave", "").strip()
        local        = request.form.get("local", "").strip()

        if not numero_chave or not local:
            flash("Preencha o número da chave e o local.", "warning")
            return redirect(url_for("chaves.meu_claviculario"))

        if not site:
            flash("Seu usuário não tem site configurado. Contate o administrador.", "danger")
            return redirect(url_for("chaves.meu_claviculario"))

        try:
            nova = ClavicularioChave(
                numero_chave = numero_chave,
                local        = local,
                site         = site,
                criador_id   = session.get("user_id"),
                criador_nome = session.get("user_nome", ""),
                ativa        = True,
            )
            _db.session.add(nova)
            _db.session.commit()
            flash("Chave adicionada ao claviculário com sucesso!", "success")
        except Exception as e:
            _db.session.rollback()
            flash(f"Erro ao salvar a chave: {e}", "danger")

        return redirect(url_for("chaves.meu_claviculario"))

    chaves = (ClavicularioChave.query
              .filter_by(site=site, ativa=True)
              .order_by(ClavicularioChave.id)
              .all())

    agora           = datetime.now()
    tres_dias_atras = agora - timedelta(days=3)
    em_uso = disponiveis = mais_de_3_dias = 0

    # ── UMA query para todas as retiradas ativas (elimina N+1) ───────────────
    chave_ids = [c.id for c in chaves]
    retiradas_map = {}
    if chave_ids:
        rows = (ClavicularioRetirada.query
                .filter(ClavicularioRetirada.chave_id.in_(chave_ids))
                .filter_by(status="EM USO")
                .all())
        retiradas_map = {r.chave_id: r for r in rows}

    for chave in chaves:
        ret = retiradas_map.get(chave.id)
        chave._retirada = ret
        if ret:
            em_uso += 1
            if ret.data_retirada <= tres_dias_atras:
                mais_de_3_dias += 1
        else:
            disponiveis += 1

    resumo = {
        "em_uso":        em_uso,
        "disponiveis":   disponiveis,
        "mais_de_3_dias": mais_de_3_dias,
    }
    return render_template(
        "chaves/meu_claviculario.html",
        chaves=chaves, resumo=resumo, site=site, is_admin=is_admin,
    )


@chaves_bp.route("/meu-claviculario/detalhe/<int:chave_id>")
@login_required_chaves
def detalhe_chave_clav(chave_id):
    chave    = ClavicularioChave.query.get_or_404(chave_id)
    site     = session.get("user_site", "")
    is_admin = (session.get("user_perfil") or "").upper() == "ADMIN"

    if chave.site != site and not is_admin:
        return jsonify({"erro": "Acesso negado"}), 403

    ret = (ClavicularioRetirada.query
           .filter_by(chave_id=chave.id, status="EM USO")
           .first())

    return jsonify({
        "id":                  chave.id,
        "numero_chave":        chave.numero_chave,
        "local":               chave.local,
        "site":                chave.site,
        "criador_nome":        chave.criador_nome,
        "status":              "EM USO" if ret else "DISPONÍVEL",
        "retirada_id":         ret.id                 if ret else None,
        "dono_temporario":     ret.nome_retirador     if ret else None,
        "cpf_matricula":       ret.cpf_matricula      if ret else None,
        "data_retirada":       ret.data_retirada.strftime("%d/%m/%Y %H:%M") if ret else None,
        "responsavel_entrega": ret.responsavel_entrega if ret else None,
        "comprovante_url":     url_for("chaves.comprovante_retirada", retirada_id=ret.id) if ret else None,
    })


@chaves_bp.route("/meu-claviculario/excluir/<int:chave_id>", methods=["POST"])
@login_required_chaves
@admin_required
def excluir_chave_clav(chave_id):
    chave  = ClavicularioChave.query.get_or_404(chave_id)
    em_uso = (ClavicularioRetirada.query
              .filter_by(chave_id=chave.id, status="EM USO")
              .first())
    if em_uso:
        flash("Não é possível remover uma chave que está em uso.", "danger")
        return redirect(url_for("chaves.meu_claviculario"))
    chave.ativa = False
    _db.session.commit()
    flash("Chave removida do claviculário.", "success")
    return redirect(url_for("chaves.meu_claviculario"))


# ═══════════════════════════════════════════════════════════════════════════════
# REALIZAR RETIRADA
# ═══════════════════════════════════════════════════════════════════════════════

@chaves_bp.route("/realizar-retirada", methods=["GET", "POST"])
@login_required_chaves
def realizar_retirada():
    site       = session.get("user_site", "")
    usuario_id = session.get("user_id")

    if request.method == "POST":
        chave_id = request.form.get("chave_id", type=int)
        chave    = ClavicularioChave.query.get(chave_id)

        if not chave or chave.site != site or not chave.ativa:
            flash("Chave inválida ou não pertence ao seu site.", "danger")
            return redirect(url_for("chaves.realizar_retirada"))

        # Bloqueia se ESSA chave específica já está em uso
        em_uso = (ClavicularioRetirada.query
                  .filter_by(chave_id=chave_id, status="EM USO")
                  .first())
        if em_uso:
            flash("Esta chave já está em uso por outra pessoa.", "danger")
            return redirect(url_for("chaves.realizar_retirada"))

        # Bloqueia se o mesmo CPF/matrícula já tem uma chave em uso no site
        cpf = request.form.get("cpf_matricula", "").strip()
        cpf_em_uso = (ClavicularioRetirada.query
                      .filter_by(cpf_matricula=cpf, site=site, status="EM USO")
                      .first())
        if cpf_em_uso:
            chave_cpf = ClavicularioChave.query.get(cpf_em_uso.chave_id)
            num = chave_cpf.numero_chave if chave_cpf else "?"
            flash(f"Este CPF/matrícula já possui a chave {num} em uso. Devolva-a antes de retirar outra.", "danger")
            return redirect(url_for("chaves.realizar_retirada"))

        try:
            nova = ClavicularioRetirada(
                chave_id            = chave_id,
                cpf_matricula       = request.form.get("cpf_matricula", "").strip(),
                nome_retirador      = request.form.get("nome_retirador", "").strip(),
                responsavel_entrega = session.get("user_nome", ""),
                usuario_id          = usuario_id,
                usuario_nome        = session.get("user_nome", ""),
                site                = site,
                data_retirada       = datetime.now(),
                status              = "EM USO",
                assinatura          = request.form.get("assinatura") or None,
            )
            _db.session.add(nova)
            _db.session.commit()
            flash(f"Retirada da chave {chave.numero_chave} registrada com sucesso!", "success")
            return redirect(url_for("chaves.realizar_retirada"))
        except Exception as e:
            _db.session.rollback()
            flash(f"Erro ao registrar retirada: {e}", "danger")
            return redirect(url_for("chaves.realizar_retirada"))

    # Chaves disponíveis do site (sem retirada ativa)
    chaves_site = (ClavicularioChave.query
                   .filter_by(site=site, ativa=True)
                   .order_by(ClavicularioChave.numero_chave)
                   .all())
    chaves_disponiveis = [
        c for c in chaves_site
        if not ClavicularioRetirada.query.filter_by(chave_id=c.id, status="EM USO").first()
    ]

    # Retiradas ativas do site para exibir no painel de status
    retiradas_ativas = (ClavicularioRetirada.query
                        .filter_by(site=site, status="EM USO")
                        .order_by(ClavicularioRetirada.data_retirada.desc())
                        .all())
    # ── UMA query para todas as chaves referenciadas (elimina N+1) ───────────
    ret_chave_ids = [r.chave_id for r in retiradas_ativas]
    chaves_ret_map = {}
    if ret_chave_ids:
        chaves_ret_map = {
            c.id: c for c in ClavicularioChave.query
                .filter(ClavicularioChave.id.in_(ret_chave_ids)).all()
        }
    agora_ret = datetime.now()
    for r in retiradas_ativas:
        chave_r = chaves_ret_map.get(r.chave_id)
        r._numero_chave = chave_r.numero_chave if chave_r else "?"
        r._local        = chave_r.local        if chave_r else "?"
        r._tempo        = _tempo_decorrido(r.data_retirada)
        r._alerta       = (agora_ret - r.data_retirada).days >= 3

    return render_template(
        "chaves/realizar_retirada.html",
        chaves_disponiveis=chaves_disponiveis,
        retiradas_ativas=retiradas_ativas,
        site=site,
        responsavel_nome=session.get("user_nome", ""),
    )


# ═══════════════════════════════════════════════════════════════════════════════
# REALIZAR DEVOLUÇÃO
# ═══════════════════════════════════════════════════════════════════════════════

@chaves_bp.route("/realizar-devolucao")
@login_required_chaves
def realizar_devolucao():
    site = session.get("user_site", "")
    retiradas = (ClavicularioRetirada.query
                 .filter_by(site=site, status="EM USO")
                 .order_by(ClavicularioRetirada.data_retirada.asc())
                 .all())

    agora = datetime.now()
    for r in retiradas:
        chave        = ClavicularioChave.query.get(r.chave_id)
        r._numero_chave = chave.numero_chave if chave else "?"
        r._local        = chave.local        if chave else "?"
        r._tempo        = _tempo_decorrido(r.data_retirada)
        r._alerta       = (agora - r.data_retirada).days >= 3

    return render_template("chaves/realizar_devolucao.html", retiradas=retiradas, site=site)


@chaves_bp.route("/realizar-devolucao/devolver/<int:retirada_id>", methods=["POST"])
@login_required_chaves
def devolver_chave_clav(retirada_id):
    retirada = ClavicularioRetirada.query.get_or_404(retirada_id)
    if retirada.status == "DEVOLVIDA":
        flash("Esta chave já foi devolvida.", "warning")
        return redirect(url_for("chaves.realizar_devolucao"))
    retirada.status         = "DEVOLVIDA"
    retirada.data_devolucao = datetime.now()
    _db.session.commit()
    flash("Chave devolvida com sucesso!", "success")
    return redirect(url_for("chaves.realizar_devolucao"))


@chaves_bp.route("/historico/export/excel")
@login_required_chaves
def exportar_historico_excel():
    """Exporta o histórico completo de retiradas (com filtros opcionais de data e status)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    site      = session.get("user_site", "")
    is_admin  = (session.get("user_perfil") or "").upper() == "ADMIN"

    # Filtros via query string
    data_ini  = request.args.get("data_ini", "").strip()
    data_fim  = request.args.get("data_fim", "").strip()
    status_f  = request.args.get("status",   "").strip().upper()  # EM USO | DEVOLVIDA | "" = todos

    query = ClavicularioRetirada.query
    if not is_admin:
        query = query.filter_by(site=site)
    else:
        site_f = request.args.get("site", site).strip()
        if site_f:
            query = query.filter_by(site=site_f)

    if status_f in ("EM USO", "DEVOLVIDA"):
        query = query.filter_by(status=status_f)

    if data_ini:
        try:
            dt_ini = datetime.strptime(data_ini, "%Y-%m-%d")
            query  = query.filter(ClavicularioRetirada.data_retirada >= dt_ini)
        except ValueError:
            pass
    if data_fim:
        try:
            dt_fim = datetime.strptime(data_fim, "%Y-%m-%d")
            dt_fim = dt_fim.replace(hour=23, minute=59, second=59)
            query  = query.filter(ClavicularioRetirada.data_retirada <= dt_fim)
        except ValueError:
            pass

    retiradas = query.order_by(ClavicularioRetirada.data_retirada.desc()).all()

    # Pré-carrega chaves para evitar N+1
    chaves_map = {}
    for r in retiradas:
        if r.chave_id not in chaves_map:
            chaves_map[r.chave_id] = ClavicularioChave.query.get(r.chave_id)

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Histórico de Retiradas"
    thin = Side(style="thin", color="D1D5DB")

    # Título
    ws.merge_cells("A1:K1")
    ws["A1"] = f"HISTÓRICO DE RETIRADAS DE CHAVES — {site}"
    ws["A1"].font      = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="D40511")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Subtítulo / metadados
    periodo = ""
    if data_ini or data_fim:
        periodo = f" | Período: {data_ini or '—'} a {data_fim or '—'}"
    filtro_status = f" | Status: {status_f}" if status_f else " | Status: Todos"
    ws.merge_cells("A2:K2")
    ws["A2"] = (f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                f" | Usuário: {session.get('user_nome', 'Sistema')}"
                f"{periodo}{filtro_status}")
    ws["A2"].font      = Font(size=10, bold=True, color="111827")
    ws["A2"].fill      = PatternFill("solid", fgColor="FFCC00")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Linha de totais resumo
    total_todos      = len(retiradas)
    total_em_uso     = sum(1 for r in retiradas if r.status == "EM USO")
    total_devolvidas = sum(1 for r in retiradas if r.status == "DEVOLVIDA")
    ws.merge_cells("A3:K3")
    ws["A3"] = (f"Total de registros: {total_todos}"
                f"  |  Em Uso: {total_em_uso}"
                f"  |  Devolvidas: {total_devolvidas}")
    ws["A3"].font      = Font(size=10, bold=True, color="374151")
    ws["A3"].fill      = PatternFill("solid", fgColor="F3F4F6")
    ws["A3"].alignment = Alignment(horizontal="center")

    # Cabeçalho das colunas
    headers = [
        "ID", "Site", "Nº Chave", "Local", "Nome do Retirador",
        "CPF / Matrícula", "Resp. Entrega", "Data Retirada",
        "Data Devolução", "Dias em Uso", "Status",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=col, value=h)
        c.font      = Font(bold=True, color="111827", size=10)
        c.fill      = PatternFill("solid", fgColor="E5E7EB")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws.row_dimensions[5].height = 22

    # Dados
    agora = datetime.now()
    fill_em_uso    = PatternFill("solid", fgColor="FEF3C7")   # amarelo claro
    fill_devolvida = PatternFill("solid", fgColor="D1FAE5")   # verde claro
    fill_alerta    = PatternFill("solid", fgColor="FEE2E2")   # vermelho claro

    for row_idx, r in enumerate(retiradas, start=6):
        chave = chaves_map.get(r.chave_id)
        dias  = (r.data_devolucao or agora) - r.data_retirada
        dias_n = dias.days

        row_data = [
            r.id,
            r.site,
            chave.numero_chave if chave else "?",
            chave.local        if chave else "?",
            r.nome_retirador   or "—",
            r.cpf_matricula,
            r.responsavel_entrega,
            r.data_retirada.strftime("%d/%m/%Y %H:%M"),
            r.data_devolucao.strftime("%d/%m/%Y %H:%M") if r.data_devolucao else "—",
            dias_n,
            r.status,
        ]
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border    = Border(top=thin, left=thin, right=thin, bottom=thin)
            c.alignment = Alignment(vertical="center")

        # Colorir a linha conforme status
        if r.status == "EM USO" and dias_n >= 3:
            row_fill = fill_alerta
        elif r.status == "EM USO":
            row_fill = fill_em_uso
        else:
            row_fill = fill_devolvida

        for col in range(1, 12):
            ws.cell(row=row_idx, column=col).fill = row_fill

    # Larguras
    for col, w in zip("ABCDEFGHIJK", [8, 14, 12, 24, 28, 18, 28, 20, 20, 12, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"

    # Legenda de cores
    row_leg = len(retiradas) + 8
    ws.cell(row=row_leg, column=1, value="Legenda:").font = Font(bold=True, size=9)
    leg_items = [
        ("DEVOLVIDA", "D1FAE5", "Chave devolvida normalmente"),
        ("EM USO (< 3 dias)", "FEF3C7", "Chave em uso dentro do prazo"),
        ("EM USO (≥ 3 dias)", "FEE2E2", "Chave em uso acima de 3 dias — requer atenção"),
    ]
    for i, (lbl, cor, desc) in enumerate(leg_items, start=1):
        c_cor  = ws.cell(row=row_leg + i, column=1, value="")
        c_cor.fill = PatternFill("solid", fgColor=cor)
        c_cor.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        c_txt = ws.cell(row=row_leg + i, column=2, value=f"{lbl} — {desc}")
        c_txt.font = Font(size=9, color="374151")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    nome_arq = f"historico_chaves_{site}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output, as_attachment=True, download_name=nome_arq,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@chaves_bp.route("/realizar-devolucao/export/excel")
@login_required_chaves
def export_devolucao_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    site      = session.get("user_site", "")
    retiradas = (ClavicularioRetirada.query
                 .filter_by(site=site, status="EM USO")
                 .order_by(ClavicularioRetirada.data_retirada.asc())
                 .all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Chaves em Uso"

    thin  = Side(style="thin", color="D1D5DB")

    # ── Cabeçalho ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    ws["A1"] = f"CHAVES EM USO — {site}"
    ws["A1"].font      = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="D40511")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:H2")
    ws["A2"] = (f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} | "
                f"Usuário: {session.get('user_nome', 'Sistema')}")
    ws["A2"].font      = Font(size=10, bold=True, color="111827")
    ws["A2"].fill      = PatternFill("solid", fgColor="FFCC00")
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Linha de headers ──────────────────────────────────────────────────────
    headers = ["Nº Chave", "Local", "CPF/Matrícula", "Usuário",
               "Responsável Entrega", "Data Retirada", "Dias em Uso", "Alerta"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font      = Font(bold=True, color="111827")
        c.fill      = PatternFill("solid", fgColor="F3F4F6")
        c.alignment = Alignment(horizontal="center")
        c.border    = Border(top=thin, left=thin, right=thin, bottom=thin)

    # ── Dados ─────────────────────────────────────────────────────────────────
    agora = datetime.now()
    for row_idx, r in enumerate(retiradas, start=5):
        chave = ClavicularioChave.query.get(r.chave_id)
        dias  = (agora - r.data_retirada).days
        row_data = [
            chave.numero_chave if chave else "?",
            chave.local        if chave else "?",
            r.cpf_matricula,
            r.usuario_nome,
            r.responsavel_entrega,
            r.data_retirada.strftime("%d/%m/%Y %H:%M"),
            dias,
            "ATENÇÃO" if dias >= 3 else "OK",
        ]
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border    = Border(top=thin, left=thin, right=thin, bottom=thin)
            c.alignment = Alignment(vertical="center")
        if dias >= 3:
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="FEE2E2")

    for col, w in zip("ABCDEFGH", [14, 24, 18, 28, 28, 20, 12, 12]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"chaves_em_uso_{site}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output, as_attachment=True, download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ═══════════════════════════════════════════════════════════════════════════════
# COMPROVANTE DE RETIRADA — PDF
# ═══════════════════════════════════════════════════════════════════════════════

@chaves_bp.route("/comprovante/<int:retirada_id>")
@login_required_chaves
def comprovante_retirada(retirada_id):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, Image as RLImage,
    )

    retirada = ClavicularioRetirada.query.get_or_404(retirada_id)
    chave    = ClavicularioChave.query.get(retirada.chave_id)

    output = BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm,
        topMargin=16*mm, bottomMargin=18*mm,
    )

    W = A4[0] - 36*mm   # largura útil
    RED  = colors.HexColor("#d40511")
    DARK = colors.HexColor("#111827")
    GRAY = colors.HexColor("#6b7280")
    LGRAY= colors.HexColor("#f3f4f6")
    YELL = colors.HexColor("#ffcc00")

    styles = getSampleStyleSheet()
    h1  = ParagraphStyle("h1",  fontName="Helvetica-Bold", fontSize=22, textColor=colors.white,
                          alignment=TA_CENTER, leading=26)
    sub = ParagraphStyle("sub", fontName="Helvetica",      fontSize=10, textColor=colors.white,
                          alignment=TA_CENTER, leading=14)
    lbl = ParagraphStyle("lbl", fontName="Helvetica-Bold", fontSize=8,  textColor=GRAY,
                          spaceAfter=2, leading=10)
    val = ParagraphStyle("val", fontName="Helvetica-Bold", fontSize=13, textColor=DARK,
                          leading=16)
    cap = ParagraphStyle("cap", fontName="Helvetica",      fontSize=8,  textColor=GRAY,
                          alignment=TA_CENTER, leading=10)
    obs = ParagraphStyle("obs", fontName="Helvetica",      fontSize=9,  textColor=GRAY,
                          alignment=TA_CENTER, leading=12)

    story = []

    # ── Cabeçalho vermelho ────────────────────────────────────────────────────
    header_data = [[
        Paragraph("COMPROVANTE DE RETIRADA DE CHAVE", h1),
    ]]
    header_table = Table(header_data, colWidths=[W])
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), RED),
        ("TOPPADDING",    (0,0), (-1,-1), 14),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("LEFTPADDING",   (0,0), (-1,-1), 12),
        ("RIGHTPADDING",  (0,0), (-1,-1), 12),
        ("ROUNDEDCORNERS", [8]),
    ]))
    story.append(header_table)

    sub_data = [[Paragraph(
        f"DHL SECURITY &nbsp;|&nbsp; {retirada.site} &nbsp;|&nbsp; "
        f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", sub,
    )]]
    sub_table = Table(sub_data, colWidths=[W])
    sub_table.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), DARK),
        ("TOPPADDING",    (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING",   (0,0), (-1,-1), 12),
        ("RIGHTPADDING",  (0,0), (-1,-1), 12),
    ]))
    story.append(sub_table)
    story.append(Spacer(1, 10*mm))

    # ── Dados da chave ────────────────────────────────────────────────────────
    def campo(rotulo, texto):
        return [Paragraph(rotulo, lbl), Paragraph(str(texto) if texto else "—", val)]

    chave_num  = chave.numero_chave if chave else "?"
    chave_local = chave.local       if chave else "?"

    chave_data = [
        [campo("Nº DA CHAVE", chave_num),   campo("LOCAL / ARMÁRIO", chave_local)],
        [campo("SITE",         retirada.site), campo("DATA E HORA DA RETIRADA",
               retirada.data_retirada.strftime("%d/%m/%Y  %H:%M"))],
    ]

    for row in chave_data:
        t = Table([[row[0][0], row[1][0]], [row[0][1], row[1][1]]], colWidths=[W/2, W/2])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), LGRAY),
            ("TOPPADDING",    (0,0), (-1,-1), 8),
            ("BOTTOMPADDING", (0,0), (-1,-1), 8),
            ("LEFTPADDING",   (0,0), (-1,-1), 10),
            ("RIGHTPADDING",  (0,0), (-1,-1), 10),
            ("LINEBELOW",     (0,0), (-1,0),  0.4, colors.HexColor("#e5e7eb")),
        ]))
        story.append(t)
        story.append(Spacer(1, 2))

    story.append(Spacer(1, 6*mm))
    story.append(HRFlowable(width=W, thickness=1, color=colors.HexColor("#e5e7eb")))
    story.append(Spacer(1, 6*mm))

    # ── Dados do Retirador ────────────────────────────────────────────────────
    ret_data = Table([
        [Paragraph("DADOS DO RETIRADOR", ParagraphStyle("sec", fontName="Helvetica-Bold",
                   fontSize=11, textColor=RED, leading=14))],
    ], colWidths=[W])
    ret_data.setStyle(TableStyle([
        ("LINEBELOW", (0,0), (-1,-1), 1.5, RED),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    story.append(ret_data)
    story.append(Spacer(1, 4*mm))

    r1 = Table([
        [campo("NOME DO RETIRADOR", retirada.nome_retirador or "—"),
         campo("CPF / MATRÍCULA",   retirada.cpf_matricula)],
    ], colWidths=[W*0.55, W*0.45])
    r1.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), LGRAY),
        ("TOPPADDING",    (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING",   (0,0), (-1,-1), 10),
        ("RIGHTPADDING",  (0,0), (-1,-1), 10),
    ]))
    story.append(r1)
    story.append(Spacer(1, 2))

    r2 = Table([
        [campo("RESPONSÁVEL PELA ENTREGA (OPERADOR)", retirada.responsavel_entrega)],
    ], colWidths=[W])
    r2.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), LGRAY),
        ("TOPPADDING",    (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING",   (0,0), (-1,-1), 10),
        ("RIGHTPADDING",  (0,0), (-1,-1), 10),
    ]))
    story.append(r2)
    story.append(Spacer(1, 8*mm))

    # ── Assinatura ────────────────────────────────────────────────────────────
    story.append(HRFlowable(width=W, thickness=1, color=colors.HexColor("#e5e7eb")))
    story.append(Spacer(1, 6*mm))

    sig_section = Table([
        [Paragraph("ASSINATURA DO RETIRADOR", ParagraphStyle("sec2", fontName="Helvetica-Bold",
                   fontSize=11, textColor=RED, leading=14))],
    ], colWidths=[W])
    sig_section.setStyle(TableStyle([
        ("LINEBELOW", (0,0), (-1,-1), 1.5, RED),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    story.append(sig_section)
    story.append(Spacer(1, 5*mm))

    if retirada.assinatura and retirada.assinatura.startswith("data:image"):
        # Decodifica base64 do canvas
        try:
            header, b64data = retirada.assinatura.split(",", 1)
            img_bytes = base64.b64decode(b64data)
            img_buf = BytesIO(img_bytes)
            sig_img = RLImage(img_buf, width=120*mm, height=38*mm)
            sig_img.hAlign = "LEFT"
            sig_box = Table([[sig_img]], colWidths=[W])
            sig_box.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,-1), colors.white),
                ("BOX",        (0,0), (-1,-1), 0.5, colors.HexColor("#d1d5db")),
                ("TOPPADDING",    (0,0), (-1,-1), 6),
                ("BOTTOMPADDING", (0,0), (-1,-1), 6),
                ("LEFTPADDING",   (0,0), (-1,-1), 8),
                ("RIGHTPADDING",  (0,0), (-1,-1), 8),
                ("ALIGN",      (0,0), (-1,-1), "CENTER"),
            ]))
            story.append(sig_box)
        except Exception:
            story.append(_caixa_assinatura_vazia(W, obs))
    else:
        story.append(_caixa_assinatura_vazia(W, obs))

    story.append(Spacer(1, 10*mm))

    # ── Rodapé ────────────────────────────────────────────────────────────────
    story.append(HRFlowable(width=W, thickness=0.5, color=colors.HexColor("#e5e7eb")))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        f"Comprovante Nº {retirada.id:05d} &nbsp;|&nbsp; DHL Security &nbsp;|&nbsp; "
        f"Documento gerado automaticamente pelo sistema CCTV Control Panel",
        ParagraphStyle("rod", fontName="Helvetica", fontSize=8, textColor=GRAY,
                       alignment=TA_CENTER, leading=10),
    ))

    doc.build(story)
    output.seek(0)
    fname = f"comprovante_retirada_{retirada.id:05d}.pdf"
    return send_file(output, as_attachment=True, download_name=fname,
                     mimetype="application/pdf")


def _caixa_assinatura_vazia(W, obs_style):
    """Retorna uma caixa tracejada para assinatura quando não há canvas."""
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    box = Table([[""]], colWidths=[W], rowHeights=[42*mm])
    box.setStyle(TableStyle([
        ("BOX",           (0,0), (-1,-1), 0.8, colors.HexColor("#cbd5e1")),
        ("BACKGROUND",    (0,0), (-1,-1), colors.HexColor("#f8fafc")),
        ("LINEBELOW",     (0,0), (-1,-1), 1.5, colors.HexColor("#94a3b8")),
    ]))
    return box
